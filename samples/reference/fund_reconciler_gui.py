#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基金资产与净值核对工具 - GUI版本 v2.1

功能：
    图形界面操作，支持文件选择、进度条展示、自定义输出路径
    
作者：技术部
版本：2.1.0
日期：2026-07-16
"""

import os
import sys
import json
import threading
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# GUI库
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except ImportError:
    print("错误：需要安装tkinter，请安装Python的GUI组件")
    sys.exit(1)


# =============================================================================
# 配置常量
# =============================================================================

DEFAULT_RENAME_MAP = {
    '7001': '6001', '7002': '7001',
    'AZ0001': '6101', 'AZ0003': '6103', 'AZ0004': '6104',
    'AZ0006': '6106', 'AZ0007': '6107', 'AZ0008': '6108',
    'AZ0010': '6110', 'AZ0011': '6111', 'AZ0012': '6112',
    'AZ0015': '6115', 'AZ0016': '6116', 'AZ0017': '6117',
    'AZ0018': '6118', 'AZ0019': '6119', 'AZ0020': '6120',
    'AZ0022': '6122', 'AZ0023': '6123', 'AZ0024': '6124',
    'AZ0026': '6126',
}

DEFAULT_BULK_PRODUCTS = {
    'AZ0206', 'AZ0205', 'AZ0221', 'AZ0230', 'AZ0130',
    'AZ0160', 'AZ0208', 'AZ0236', 'AZ0210', 'AZ0151',
    'AZ0207',
}

DEFAULT_DIFF_THRESHOLD = 1.0
DEFAULT_SIMILARITY_THRESHOLD = 0.5


# =============================================================================
# 核心核对逻辑
# =============================================================================

def similarity(a, b):
    return SequenceMatcher(None, str(a), str(b)).ratio()


class FundAssetReconciler:
    """基金资产与净值核对器（带进度回调）"""
    
    def __init__(self, progress_callback=None, log_callback=None):
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.rename_map = DEFAULT_RENAME_MAP.copy()
        self.bulk_products = DEFAULT_BULK_PRODUCTS.copy()
        self.diff_threshold = DEFAULT_DIFF_THRESHOLD
        self.similarity_threshold = DEFAULT_SIMILARITY_THRESHOLD
        
    def log(self, message):
        if self.log_callback:
            self.log_callback(message)
            
    def update_progress(self, value, message=""):
        if self.progress_callback:
            self.progress_callback(value, message)
    
    def validate_output_path(self, output_path):
        """验证并修正输出路径"""
        # 去除首尾空格
        output_path = output_path.strip()
        
        # 如果路径是目录，添加默认文件名
        if os.path.isdir(output_path):
            default_name = f"基金资产与净值核对结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(output_path, default_name)
            self.log(f"检测到输出路径为目录，自动添加文件名: {default_name}")
        
        # 确保有.xlsx扩展名
        if not output_path.lower().endswith('.xlsx'):
            output_path += '.xlsx'
        
        # 检查目录是否存在，不存在则创建
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                self.log(f"创建输出目录: {output_dir}")
            except Exception as e:
                raise ValueError(f"无法创建输出目录 '{output_dir}': {str(e)}")
        
        # 检查目录是否可写
        if output_dir and os.path.exists(output_dir):
            try:
                test_file = os.path.join(output_dir, '.write_test')
                with open(test_file, 'w') as f:
                    f.write('test')
                os.remove(test_file)
            except Exception as e:
                raise ValueError(f"输出目录 '{output_dir}' 没有写入权限: {str(e)}")
        
        return output_path
    
    def reconcile(self, fund_path, netvalue_path, output_path):
        """执行核对流程"""
        # 验证并修正输出路径
        output_path = self.validate_output_path(output_path)
        self.log(f"最终输出路径: {output_path}")
        
        self.update_progress(5, "正在读取基金资产表...")
        df_fund = pd.read_excel(fund_path, header=0)
        self.log(f"基金资产表记录数: {len(df_fund)}, 列数: {len(df_fund.columns)}")
        
        self.update_progress(10, "正在读取净值查询表...")
        df_net = pd.read_excel(netvalue_path, header=0)
        self.log(f"净值查询表记录数: {len(df_net)}, 列数: {len(df_net.columns)}")
        
        # 验证列数
        if len(df_fund.columns) < 28:
            raise ValueError(
                f"基金资产表列数不足！当前{len(df_fund.columns)}列，需要至少28列。\n"
                f"请确认：\n"
                f"1. 选择的文件是否正确（基金资产表应该有'基金编号'、'基金名称'等列）\n"
                f"2. 是否将净值查询表误选为基金资产表"
            )
        
        if len(df_net.columns) < 9:
            raise ValueError(
                f"净值查询表列数不足！当前{len(df_net.columns)}列，需要至少9列。\n"
                f"请确认：\n"
                f"1. 选择的文件是否正确（净值查询表应该有'信托计划代码'、'资产净值'等列）\n"
                f"2. 是否将基金资产表误选为净值查询表"
            )
        
        # 检测文件是否可能选反
        if len(df_fund.columns) < 15 and len(df_net.columns) > 25:
            self.log("⚠️ 警告：检测到文件列数异常，可能文件选反了！")
            self.log(f"  基金资产表列数: {len(df_fund.columns)} (通常>25)")
            self.log(f"  净值查询表列数: {len(df_net.columns)} (通常~9)")
        
        self.update_progress(15, "正在提取关键列...")
        df_fund_key = df_fund.iloc[:, [2, 3, 11, 27, 19]].copy()
        df_fund_key.columns = ['组合编号', '组合名称', '基金资产净值', '基金总资产', '信托计划代码']
        
        df_net_key = df_net.iloc[:, [8, 4, 2, 0]].copy()
        df_net_key.columns = ['信托计划代码', '查询总资产', '查询资产净值', '产品名称']
        
        self.update_progress(20, "正在应用重命名映射...")
        df_net_key['原始代码'] = df_net_key['信托计划代码'].astype(str)
        df_net_key['信托计划代码'] = df_net_key['原始代码'].replace(self.rename_map)
        df_net_key['映射记录'] = df_net_key['原始代码'] != df_net_key['信托计划代码']
        
        self.update_progress(25, "正在映射去重...")
        grouped = df_net_key.groupby('信托计划代码', sort=False)
        filtered_rows = []
        for code, group in grouped:
            if group['映射记录'].any() and not group['映射记录'].all():
                filtered_rows.append(group[group['映射记录']])
            else:
                filtered_rows.append(group)
        df_net_key = pd.concat(filtered_rows, ignore_index=True).drop(columns=['原始代码', '映射记录'])
        
        self.update_progress(30, "正在过滤I9记录...")
        df_net_key['含I9'] = df_net_key['产品名称'].str.contains('I9', case=False, na=False)
        grouped = df_net_key.groupby('信托计划代码', sort=False)
        filtered_rows = []
        for code, group in grouped:
            if len(group) > 1 and not group['含I9'].all():
                filtered_rows.append(group[~group['含I9']])
            else:
                filtered_rows.append(group)
        df_net_key = pd.concat(filtered_rows, ignore_index=True).drop(columns=['含I9'])
        
        self.update_progress(40, "正在精确匹配...")
        merged = pd.merge(df_fund_key, df_net_key, on='信托计划代码', how='left')
        merged['匹配方式'] = np.where(merged['查询总资产'].notna(), '精确', '未匹配')
        merged['匹配相似度'] = None
        exact_count = (merged['匹配方式'] == '精确').sum()
        self.log(f"精确匹配: {exact_count}条")
        
        self.update_progress(50, "正在模糊匹配...")
        unmatched = merged[merged['查询总资产'].isna()].copy()
        matched_codes = set(merged[merged['查询总资产'].notna()]['信托计划代码'].dropna().unique())
        available_net = df_net_key[~df_net_key['信托计划代码'].isin(matched_codes)].copy()
        
        used_net_indices = set()
        fuzzy_count = 0
        total_unmatched = len(unmatched)
        
        for idx, (idx1, row1) in enumerate(unmatched.iterrows()):
            best_score = 0
            best_idx2 = None
            best_row2 = None
            
            for idx2, row2 in available_net.iterrows():
                if idx2 in used_net_indices:
                    continue
                score = similarity(str(row1['组合名称']), str(row2['产品名称']))
                if score > best_score:
                    best_score = score
                    best_idx2 = idx2
                    best_row2 = row2
            
            if best_score >= self.similarity_threshold and best_idx2 is not None:
                merged.loc[idx1, '查询总资产'] = best_row2['查询总资产']
                merged.loc[idx1, '查询资产净值'] = best_row2['查询资产净值']
                merged.loc[idx1, '产品名称'] = best_row2['产品名称']
                merged.loc[idx1, '匹配方式'] = '模糊'
                merged.loc[idx1, '匹配相似度'] = round(best_score, 4)
                used_net_indices.add(best_idx2)
                fuzzy_count += 1
            
            if idx % 5 == 0:
                progress = 50 + int((idx / max(total_unmatched, 1)) * 20)
                self.update_progress(progress, f"正在模糊匹配... ({idx}/{total_unmatched})")
        
        self.log(f"模糊匹配: {fuzzy_count}条")
        
        self.update_progress(70, "正在计算差异...")
        merged['总资产差值'] = merged['基金总资产'] - merged['查询总资产']
        merged['总资产差值百分比'] = np.where(
            merged['查询总资产'].notna() & (merged['查询总资产'] != 0),
            (merged['基金总资产'] - merged['查询总资产']) / merged['查询总资产'] * 100,
            np.nan
        )
        merged['资产净值差值'] = merged['基金资产净值'] - merged['查询资产净值']
        merged['资产净值差值百分比'] = np.where(
            merged['查询资产净值'].notna() & (merged['查询资产净值'] != 0),
            (merged['基金资产净值'] - merged['查询资产净值']) / merged['查询资产净值'] * 100,
            np.nan
        )
        
        self.update_progress(80, "正在排序...")
        merged['最大差异百分比'] = merged[['总资产差值百分比', '资产净值差值百分比']].abs().max(axis=1)
        merged['未匹配'] = merged['查询总资产'].isna()
        merged = merged.sort_values(by=['未匹配', '最大差异百分比'], ascending=[True, False])
        merged = merged.drop(columns=['最大差异百分比', '未匹配'])
        merged = merged.reset_index(drop=True)
        
        merged['是否大宗'] = merged['信托计划代码'].isin(self.bulk_products)
        
        self.update_progress(85, "正在生成Excel...")
        self.generate_excel(merged, output_path)
        
        self.update_progress(100, "核对完成！")
        
        # 统计
        stats = {
            '总记录数': len(merged),
            '精确匹配': int((merged['匹配方式'] == '精确').sum()),
            '模糊匹配': int((merged['匹配方式'] == '模糊').sum()),
            '未匹配': int((merged['匹配方式'] == '未匹配').sum()),
            '大宗产品数': int(merged['是否大宗'].sum()),
        }
        
        diff_df = merged[(merged['匹配方式'] != '未匹配') & (~merged['是否大宗'])].copy()
        diff_df['最大差异百分比'] = diff_df[['总资产差值百分比', '资产净值差值百分比']].abs().max(axis=1)
        stats['差异大于1%数量'] = int((diff_df['最大差异百分比'] > self.diff_threshold).sum())
        
        return stats
    
    def generate_excel(self, df, output_path):
        """生成Excel报告"""
        wb = Workbook()
        ws = wb.active
        ws.title = "核对结果"
        
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        header_fill = PatternFill(start_color='B4C7DC', end_color='B4C7DC', fill_type='solid')
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ['组合编号', '组合名称', '信托计划代码', '产品名称', '匹配方式', '匹配相似度',
                   '基金总资产', '查询总资产', '总资产差值', '总资产差值百分比(%)',
                   '基金资产净值', '查询资产净值', '资产净值差值', '资产净值差值百分比(%)', '差异原因']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = header_fill
        
        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2
            is_bulk = row.get('是否大宗', False)
            is_unmatched = row['匹配方式'] == '未匹配'
            is_diff = False
            
            if not is_unmatched and not is_bulk:
                total_diff = abs(row['总资产差值百分比']) if pd.notna(row['总资产差值百分比']) else 0
                nav_diff = abs(row['资产净值差值百分比']) if pd.notna(row['资产净值差值百分比']) else 0
                is_diff = total_diff > self.diff_threshold or nav_diff > self.diff_threshold
            
            reason = '大宗产品无需核对' if is_bulk else ''
            
            values = [
                row['组合编号'], row['组合名称'], row['信托计划代码'], row['产品名称'],
                row['匹配方式'], row['匹配相似度'],
                row['基金总资产'], row['查询总资产'], row['总资产差值'], row['总资产差值百分比'],
                row['基金资产净值'], row['查询资产净值'], row['资产净值差值'], row['资产净值差值百分比'],
                reason
            ]
            
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=col, value=val)
                cell.border = thin_border
                if col in [7, 8, 9, 11, 12, 13]:
                    cell.number_format = '#,##0.00'
                elif col in [10, 14] and val is not None:
                    cell.number_format = '0.00'
                
                if is_bulk:
                    cell.fill = orange_fill
                elif is_unmatched:
                    cell.fill = yellow_fill
                elif is_diff:
                    cell.fill = red_fill
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 22
        ws.column_dimensions['O'].width = 20
        
        ws.freeze_panes = 'A2'
        wb.save(output_path)
        self.log(f"Excel报告已保存: {output_path}")


# =============================================================================
# GUI界面
# =============================================================================

class ReconcilerGUI:
    """基金资产核对工具图形界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("基金资产与净值核对工具 v2.1")
        self.root.geometry("700x520")
        self.root.resizable(False, False)
        
        # 设置窗口居中
        self.center_window()
        
        # 样式
        self.style = ttk.Style()
        self.style.configure('TButton', font=('微软雅黑', 10))
        self.style.configure('TLabel', font=('微软雅黑', 10))
        self.style.configure('TEntry', font=('微软雅黑', 10))
        
        self.create_widgets()
        
        # 文件路径变量
        self.fund_path = tk.StringVar()
        self.netvalue_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        # 设置默认输出路径（桌面）
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        default_output = os.path.join(desktop, 
                                      f"基金资产与净值核对结果_{datetime.now().strftime('%Y%m%d')}.xlsx")
        self.output_path.set(default_output)
    
    def center_window(self):
        """窗口居中"""
        self.root.update_idletasks()
        width = 700
        height = 520
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 标题
        title_label = ttk.Label(main_frame, text="基金资产与净值核对工具", 
                                font=('微软雅黑', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # 基金资产表选择
        ttk.Label(main_frame, text="基金资产表:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.fund_entry = ttk.Entry(main_frame, width=50, font=('微软雅黑', 9))
        self.fund_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(main_frame, text="浏览...", command=self.browse_fund).grid(row=1, column=2, padx=5)
        
        # 净值查询表选择
        ttk.Label(main_frame, text="净值查询表:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.netvalue_entry = ttk.Entry(main_frame, width=50, font=('微软雅黑', 9))
        self.netvalue_entry.grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(main_frame, text="浏览...", command=self.browse_netvalue).grid(row=2, column=2, padx=5)
        
        # 输出路径选择
        ttk.Label(main_frame, text="输出路径:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.output_entry = ttk.Entry(main_frame, width=50, font=('微软雅黑', 9))
        self.output_entry.grid(row=3, column=1, padx=5, pady=5)
        ttk.Button(main_frame, text="浏览...", command=self.browse_output).grid(row=3, column=2, padx=5)
        
        # 提示标签
        tip_label = ttk.Label(main_frame, text="提示：输出路径可以是具体文件路径，也可以是目录（将自动生成文件名）", 
                              font=('微软雅黑', 8), foreground='gray')
        tip_label.grid(row=4, column=1, sticky=tk.W)
        
        # 分隔线
        ttk.Separator(main_frame, orient='horizontal').grid(row=5, column=0, columnspan=3, 
                                                             sticky=(tk.W, tk.E), pady=15)
        
        # 执行按钮
        self.run_button = ttk.Button(main_frame, text="开始核对", command=self.run_reconciliation,
                                     style='Accent.TButton')
        self.run_button.grid(row=6, column=0, columnspan=3, pady=10)
        self.style.configure('Accent.TButton', font=('微软雅黑', 12, 'bold'))
        
        # 进度条
        ttk.Label(main_frame, text="执行进度:").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                            maximum=100, length=500, mode='determinate')
        self.progress_bar.grid(row=7, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # 进度文字
        self.progress_label = ttk.Label(main_frame, text="就绪", font=('微软雅黑', 9))
        self.progress_label.grid(row=8, column=1, sticky=tk.W)
        
        # 日志区域
        ttk.Label(main_frame, text="执行日志:").grid(row=9, column=0, sticky=tk.W, pady=(10, 5))
        
        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=10, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.log_text = tk.Text(log_frame, width=80, height=10, font=('Consolas', 9),
                                wrap=tk.WORD, state='disabled')
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.log_text['yscrollcommand'] = scrollbar.set
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, 
                               font=('微软雅黑', 9), relief=tk.SUNKEN)
        status_bar.grid(row=11, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def browse_fund(self):
        """选择基金资产表"""
        file_path = filedialog.askopenfilename(
            title="选择基金资产表",
            filetypes=[("Excel文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.fund_entry.delete(0, tk.END)
            self.fund_entry.insert(0, file_path)
    
    def browse_netvalue(self):
        """选择净值查询表"""
        file_path = filedialog.askopenfilename(
            title="选择净值查询表",
            filetypes=[("Excel文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.netvalue_entry.delete(0, tk.END)
            self.netvalue_entry.insert(0, file_path)
    
    def browse_output(self):
        """选择输出路径"""
        # 使用 asksaveasfilename 让用户选择文件
        file_path = filedialog.asksaveasfilename(
            title="选择输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"基金资产与净值核对结果_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if file_path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, file_path)
    
    def log(self, message):
        """添加日志"""
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')
    
    def update_progress(self, value, message=""):
        """更新进度"""
        self.progress_var.set(value)
        if message:
            self.progress_label.config(text=message)
        self.root.update_idletasks()
    
    def run_reconciliation(self):
        """执行核对"""
        fund_path = self.fund_entry.get().strip()
        netvalue_path = self.netvalue_entry.get().strip()
        output_path = self.output_entry.get().strip()
        
        # 验证输入
        if not fund_path or not os.path.exists(fund_path):
            messagebox.showerror("错误", "请选择有效的基金资产表文件！")
            return
        
        if not netvalue_path or not os.path.exists(netvalue_path):
            messagebox.showerror("错误", "请选择有效的净值查询表文件！")
            return
        
        if not output_path:
            messagebox.showerror("错误", "请选择输出路径！")
            return
        
        # 禁用按钮
        self.run_button.config(state='disabled')
        self.status_var.set("正在执行核对...")
        
        # 清空日志
        self.log_text.config(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state='disabled')
        
        # 在新线程中执行核对
        thread = threading.Thread(target=self._do_reconciliation, 
                                  args=(fund_path, netvalue_path, output_path))
        thread.daemon = True
        thread.start()
    
    def _do_reconciliation(self, fund_path, netvalue_path, output_path):
        """在后台线程执行核对"""
        try:
            self.log("=" * 50)
            self.log("开始基金资产与净值核对")
            self.log("=" * 50)
            self.log(f"基金资产表: {fund_path}")
            self.log(f"净值查询表: {netvalue_path}")
            self.log(f"输出路径: {output_path}")
            
            reconciler = FundAssetReconciler(
                progress_callback=self.update_progress,
                log_callback=self.log
            )
            
            stats = reconciler.reconcile(fund_path, netvalue_path, output_path)
            
            self.log("=" * 50)
            self.log("核对完成！")
            for key, value in stats.items():
                self.log(f"  {key}: {value}")
            self.log("=" * 50)
            
            # 显示完成提示
            self.root.after(0, lambda: self._show_completion(output_path, stats))
            
        except ValueError as ve:
            # 用户输入错误（如文件选反、列数不足）
            self.log(f"❌ 验证错误: {str(ve)}")
            self.root.after(0, lambda: messagebox.showerror("文件验证错误", str(ve)))
        except PermissionError as pe:
            # 权限错误
            self.log(f"❌ 权限错误: {str(pe)}")
            self.root.after(0, lambda: messagebox.showerror("权限错误", 
                f"无法保存到指定路径，请检查：\n\n"
                f"1. 输出路径是否为文件夹而非文件\n"
                f"2. 是否有写入权限\n"
                f"3. 文件是否被其他程序占用\n\n"
                f"原始错误: {str(pe)}"))
        except Exception as e:
            self.log(f"❌ 错误: {str(e)}")
            import traceback
            error_detail = traceback.format_exc()
            self.log(f"错误详情:\n{error_detail}")
            self.root.after(0, lambda: messagebox.showerror("执行错误", f"核对过程中出现错误:\n{str(e)}\n\n请检查日志了解详情。"))
        finally:
            self.root.after(0, lambda: self.run_button.config(state='normal'))
            self.root.after(0, lambda: self.status_var.set("就绪"))
    
    def _show_completion(self, output_path, stats):
        """显示完成对话框"""
        msg = f"核对完成！\n\n"
        msg += f"总记录数: {stats['总记录数']}\n"
        msg += f"精确匹配: {stats['精确匹配']}\n"
        msg += f"模糊匹配: {stats['模糊匹配']}\n"
        msg += f"未匹配: {stats['未匹配']}\n"
        msg += f"大宗产品数: {stats['大宗产品数']}\n"
        msg += f"差异>1%数量: {stats['差异大于1%数量']}\n\n"
        msg += f"结果文件已保存:\n{output_path}"
        
        if messagebox.askyesno("核对完成", msg + "\n\n是否打开结果文件？"):
            os.startfile(output_path)


def main():
    """主函数"""
    root = tk.Tk()
    
    # 设置DPI感知
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
    
    app = ReconcilerGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
