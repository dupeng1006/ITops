# -*- coding: utf-8 -*-
"""
特殊产品清单（差异说明 note / 行填充色 color）引擎单元测试

覆盖：
    A. 默认行为一字不差（红线）：11 个默认特殊产品 → 差异原因="大宗产品无需核对"、
       填充色 FFC000、统计与黄金基线一致（16/14/1/1/2/1）；
    B. 自定义 note/color：M1 全流程后 openpyxl 读回结果 xlsx，断言该行
       "差异原因"=自定义 note、填充色=自定义 color；未配置产品不受影响；
    C. 向后兼容：set 传参（旧 bulk_products 用法）行为等同默认；
    D. SpecialProductRule 对象传参（DbRuleProvider 路径）与 # 前缀颜色归一化；
    E. 统计逻辑不动：自定义 note/color 不改变 stats。

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests\\unit\\test_special_products.py
退出码：全部通过 0；任一失败非零。

作者：技术部
版本：1.0.0
日期：2026-07-20
"""

import sys
import tempfile
from pathlib import Path

SERVER_ROOT = Path(__file__).resolve().parents[2]
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:  # noqa: BLE001
    pass

from openpyxl import load_workbook  # noqa: E402

from app.engines.fund_reconciler_base import FundAssetReconciler  # noqa: E402
from app.services.rule_service import SpecialProductRule  # noqa: E402

PROJECT_ROOT = SERVER_ROOT.parent
FUND_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "基金资产表_样本.xlsx"
NET_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "净值查询表_样本.xlsx"

# 黄金基线统计（tests/golden/expected/expected_stats.json 同值）
BASELINE_STATS = {
    '总记录数': 16, '精确匹配': 14, '模糊匹配': 1, '未匹配': 1,
    '大宗产品数': 2, '差异>1.0%数量（非大宗）': 1,
}

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def _run(special_products) -> Path:
    out = Path(tempfile.mkdtemp(prefix="special_")) / "result.xlsx"
    r = FundAssetReconciler(special_products=special_products)
    stats = r.reconcile(str(FUND_SAMPLE), str(NET_SAMPLE), str(out))
    return out, stats, r


def _row_of(ws, code: str):
    for row in ws.iter_rows(min_row=2):
        if row[2].value == code:
            return row
    return None


def test_default_behavior() -> None:
    """A. 默认行为与现状一字不差（红线）"""
    out, stats, _ = _run(None)
    check("A1 默认统计==黄金基线", stats == BASELINE_STATS, str(stats))
    ws = load_workbook(out).active
    for code in ("AZ0206", "AZ0205"):
        row = _row_of(ws, code)
        check(f"A2 默认 {code} 差异原因=大宗产品无需核对",
              row is not None and row[14].value == "大宗产品无需核对",
              str(row[14].value if row else None))
        check(f"A3 默认 {code} 填充色 FFC000",
              row is not None and row[0].fill.start_color.rgb.endswith("FFC000"),
              str(row[0].fill.start_color.rgb if row else None))


def test_custom_note_color() -> None:
    """B. 自定义 note/color 落 Excel"""
    out, stats, _ = _run({
        'AZ0206': {'note': '大额申赎差异(月末确认)', 'color': '92D050'},
        'AZ0205': {'note': None, 'color': 'FFC000'},
    })
    check("B1 自定义后统计仍==黄金基线（统计逻辑不动）",
          stats == BASELINE_STATS, str(stats))
    ws = load_workbook(out).active
    row = _row_of(ws, "AZ0206")
    check("B2 自定义差异原因写入'差异原因'列",
          row is not None and row[14].value == "大额申赎差异(月末确认)",
          str(row[14].value if row else None))
    check("B3 自定义填充色 92D050",
          row is not None and row[0].fill.start_color.rgb.endswith("92D050"),
          str(row[0].fill.start_color.rgb if row else None))
    row2 = _row_of(ws, "AZ0205")
    check("B4 未配 note 产品仍默认文案",
          row2 is not None and row2[14].value == "大宗产品无需核对",
          str(row2[14].value if row2 else None))
    # 非特殊产品行不受自定义影响（抽查首个非大宗行的差异原因列为空或非大宗文案）
    other = _row_of(ws, "AZ0001") or _row_of(ws, "7001")
    check("B5 非特殊产品行无大宗标注",
          other is None or other[14].value != "大宗产品无需核对", "n/a")


def test_backward_compat() -> None:
    """C/D. 传参兼容路径"""
    out, stats, r = _run({'AZ0206', 'AZ0205'})
    check("C1 set 传参统计==黄金基线", stats == BASELINE_STATS, str(stats))
    ws = load_workbook(out).active
    row = _row_of(ws, "AZ0206")
    check("C2 set 传参默认文案+FFC000",
          row is not None and row[14].value == "大宗产品无需核对"
          and row[0].fill.start_color.rgb.endswith("FFC000"),
          str(row[14].value if row else None))

    r2 = FundAssetReconciler(special_products={
        'AZ0206': SpecialProductRule(note='规则对象说明', color='#00b0f0'),
    })
    check("D1 SpecialProductRule 对象传参（#前缀归一化大写）",
          r2.special_products['AZ0206'] == {'note': '规则对象说明', 'color': '00B0F0'},
          str(r2.special_products['AZ0206']))
    out2, _, _ = _run({'AZ0206': SpecialProductRule(note='规则对象说明', color='00B0F0'),
                       'AZ0205': SpecialProductRule()})
    ws2 = load_workbook(out2).active
    row3 = _row_of(ws2, "AZ0206")
    check("D2 对象传参 note/color 落 Excel",
          row3 is not None and row3[14].value == '规则对象说明'
          and row3[0].fill.start_color.rgb.endswith('00B0F0'),
          str((row3[14].value, row3[0].fill.start_color.rgb) if row3 else None))


if __name__ == "__main__":
    test_default_behavior()
    test_custom_note_color()
    test_backward_compat()
    print(f"\n{'全部通过' if not failures else f'{len(failures)} 项失败'}")
    for f in failures:
        print(f"  - {f}")
    sys.exit(0 if not failures else 1)
