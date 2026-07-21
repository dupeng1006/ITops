# -*- coding: utf-8 -*-
"""
M3 黄金样本回归测试（纯脚本，无 pytest 依赖）

比对项：
    1. 统计 dict：与 expected/expected_stats.json 逐项相等；
    2. 明细 DataFrame：与 expected/expected_detail_df.pkl 逐行逐列值相等；
    3. 更新表银行间ID 列值：与预期 Excel 逐格一致（含空值保持）；
    4. 三类颜色抽样：更新表与明细表的 蓝BDD7EE/绿C6EFCE/红FFC7CE 代表行逐格一致；
    5. 说明 md 关键内容：统计数字、三类未匹配归类提示、有变动明细行数；
    6. 输入校验单测：缺关键列抛 ValueError，中文提示指明缺哪列。

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests/golden/m3/test_golden_m3.py
退出码：全部通过 0；任一不一致打印差异明细并以非零码退出。

作者：技术部
版本：1.0.0
日期：2026-07-17
"""

import json
import pickle
import sys
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# 将 server/ 加入模块搜索路径（本文件位于 server/tests/golden/m3/）
SERVER_ROOT = Path(__file__).resolve().parents[3]
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

from app.engines.m3_interbank_id import (  # noqa: E402
    MATCH_CHANGED,
    MATCH_NONE,
    MATCH_SAME,
    M3InterbankIdEngine,
)

M3_DIR = Path(__file__).resolve().parent
SAMPLE_DIR = M3_DIR / "samples"
EXPECTED_DIR = M3_DIR / "expected"
ACTUAL_DIR = M3_DIR / "actual"

FUND_SAMPLE = SAMPLE_DIR / "基金属性表_样本.xlsx"
MEMBER_SAMPLE = SAMPLE_DIR / "交易成员基本信息表_样本.csv"

EXPECTED_UPDATED = EXPECTED_DIR / "基金属性_精确匹配更新.xlsx"
EXPECTED_DETAIL = EXPECTED_DIR / "精确匹配结果明细.xlsx"
EXPECTED_NOTE = EXPECTED_DIR / "精确匹配说明.md"
EXPECTED_STATS = EXPECTED_DIR / "expected_stats.json"
EXPECTED_DF = EXPECTED_DIR / "expected_detail_df.pkl"

ACTUAL_UPDATED = ACTUAL_DIR / "基金属性_精确匹配更新.xlsx"
ACTUAL_DETAIL = ACTUAL_DIR / "精确匹配结果明细.xlsx"
ACTUAL_NOTE = ACTUAL_DIR / "精确匹配说明.md"

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def cell_rgb(ws, row: int, col: int) -> str:
    return ws.cell(row=row, column=col).fill.start_color.rgb or ""


# =============================================================================
# 检查 1：统计 dict
# =============================================================================

def check_stats(actual_stats: dict) -> None:
    with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
        expected_stats = json.load(f)
    check("统计 dict 逐项相等", actual_stats == expected_stats,
          f"预期{expected_stats} != 实际{actual_stats}")


# =============================================================================
# 检查 2：明细 DataFrame 逐行逐列
# =============================================================================

def check_detail_df(actual_df: pd.DataFrame) -> None:
    with open(EXPECTED_DF, "rb") as f:
        expected_df = pickle.load(f)
    check("明细 DataFrame 形状一致", expected_df.shape == actual_df.shape,
          f"预期{expected_df.shape} != 实际{actual_df.shape}")
    if expected_df.shape != actual_df.shape:
        return
    check("明细 DataFrame 列名一致",
          list(expected_df.columns) == list(actual_df.columns),
          f"预期{list(expected_df.columns)} != 实际{list(actual_df.columns)}")
    mismatch = []
    for col in expected_df.columns:
        for idx in range(len(expected_df)):
            e, a = expected_df.iloc[idx][col], actual_df.iloc[idx][col]
            e_na, a_na = pd.isna(e), pd.isna(a)
            if e_na or a_na:
                if not (e_na and a_na):
                    mismatch.append(f"行{idx} 列[{col}]: 预期{e!r} != 实际{a!r}")
            elif e != a:
                mismatch.append(f"行{idx} 列[{col}]: 预期{e!r} != 实际{a!r}")
    check("明细 DataFrame 逐行逐列值相等", not mismatch,
          "\n    ".join(mismatch[:20]))


# =============================================================================
# 检查 3：更新表银行间ID 列值逐格一致
# =============================================================================

def check_updated_id_column() -> None:
    wb_e = load_workbook(EXPECTED_UPDATED)
    wb_a = load_workbook(ACTUAL_UPDATED)
    ws_e, ws_a = wb_e.active, wb_a.active
    check("更新表行数一致", ws_e.max_row == ws_a.max_row,
          f"预期{ws_e.max_row} != 实际{ws_a.max_row}")
    # 银行间ID 为第 4 列（基金代码/基金全称/基金类型/银行间ID/备注）
    mismatch = []
    for r in range(2, ws_e.max_row + 1):
        e = ws_e.cell(row=r, column=4).value
        a = ws_a.cell(row=r, column=4).value
        if (e or "") != (a or ""):
            mismatch.append(f"行{r}: 预期{e!r} != 实际{a!r}")
    check("更新表银行间ID 列值逐格一致", not mismatch, "\n    ".join(mismatch[:20]))


# =============================================================================
# 检查 4：三类颜色抽样（更新表 + 明细表，预期 vs 实际逐格一致 + 色值正确）
# =============================================================================

def check_colors() -> None:
    # (文件对, 抽样行: (行号, 期望色)) 行2-5蓝, 6-10绿, 11-16红
    samples = [(2, "BDD7EE"), (5, "BDD7EE"), (6, "C6EFCE"), (10, "C6EFCE"),
               (11, "FFC7CE"), (16, "FFC7CE")]
    for label, exp_path, act_path in (
        ("更新表", EXPECTED_UPDATED, ACTUAL_UPDATED),
        ("明细表", EXPECTED_DETAIL, ACTUAL_DETAIL),
    ):
        ws_e = load_workbook(exp_path).active
        ws_a = load_workbook(act_path).active
        for row, color in samples:
            rgb_a = cell_rgb(ws_a, row, 1)
            rgb_e = cell_rgb(ws_e, row, 1)
            check(f"{label} 第{row}行颜色={color}",
                  rgb_a.endswith(color) and rgb_a == rgb_e,
                  f"实际{rgb_a} 预期{rgb_e}")


# =============================================================================
# 检查 5：说明 md 关键内容
# =============================================================================

def check_note_md() -> None:
    text = ACTUAL_NOTE.read_text(encoding="utf-8")
    expected_text = EXPECTED_NOTE.read_text(encoding="utf-8")
    check("说明 md 与预期完全一致", text == expected_text,
          "md 内容存在差异")
    for keyword in ("总记录数: 15", f"{MATCH_CHANGED}: 5", f"{MATCH_SAME}: 4",
                    f"{MATCH_NONE}: 6", "内部简称（TR_ 前缀）", "未注册产品",
                    "命名后缀不一致", "测试进取5号", "TR_测试现金管理1号"):
        check(f"说明 md 含关键内容[{keyword}]", keyword in text, keyword)
    # 有变动明细表体行数 = 5
    section = text.split("## 二、精确匹配-有变动明细")[1].split("## 三")[0]
    body_rows = [ln for ln in section.splitlines()
                 if ln.startswith("|") and "---" not in ln and "基金全称" not in ln]
    check("说明 md 有变动明细行数=5", len(body_rows) == 5, f"实际{len(body_rows)}")


# =============================================================================
# 检查 6：输入校验单测（缺列中文报错）
# =============================================================================

def check_validation() -> None:
    import tempfile
    engine = M3InterbankIdEngine()
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        # 基金属性表缺 银行间ID
        bad_fund = tmp / "缺列基金属性表.xlsx"
        pd.DataFrame({"基金代码": ["F001"], "基金全称": ["测试产品"]}).to_excel(bad_fund, index=False)
        try:
            engine.validate_input_files(bad_fund, MEMBER_SAMPLE)
            check("基金属性表缺列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("基金属性表缺列抛 ValueError", True)
            check("报错指明缺[银行间ID]列", "银行间ID" in str(e), str(e))

        # 交易成员表缺 交易成员ID
        bad_member = tmp / "缺列交易成员表.csv"
        pd.DataFrame({"交易成员全称": ["测试成员"], "机构代码": ["ORG001"]}).to_csv(
            bad_member, index=False, encoding="gbk")
        try:
            engine.validate_input_files(FUND_SAMPLE, bad_member)
            check("交易成员表缺列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("交易成员表缺列抛 ValueError", True)
            check("报错指明缺[交易成员ID]列", "交易成员ID" in str(e), str(e))


# =============================================================================
# 主流程
# =============================================================================

def main() -> int:
    print("=" * 70)
    print("M3 黄金样本回归测试（基金属性表银行间ID匹配）")
    print("=" * 70)

    for path in (FUND_SAMPLE, MEMBER_SAMPLE, EXPECTED_UPDATED, EXPECTED_DETAIL,
                 EXPECTED_NOTE, EXPECTED_STATS, EXPECTED_DF):
        if not path.exists():
            print(f"缺失文件: {path}，请先运行 make_samples_m3.py 与 build_expected_m3.py")
            return 2

    ACTUAL_DIR.mkdir(parents=True, exist_ok=True)
    engine = M3InterbankIdEngine()
    result = engine.run(
        fund_path=str(FUND_SAMPLE),
        member_path=str(MEMBER_SAMPLE),
        output_dir=str(ACTUAL_DIR),
    )
    print("-" * 70)
    print("【检查 1】统计 dict")
    check_stats(result["stats"])
    print("【检查 2】明细 DataFrame")
    check_detail_df(result["detail_df"])
    print("【检查 3】更新表银行间ID 列值")
    check_updated_id_column()
    print("【检查 4】三类颜色抽样")
    check_colors()
    print("【检查 5】说明 md 关键内容")
    check_note_md()
    print("【检查 6】输入校验单测")
    check_validation()

    print("=" * 70)
    if failures:
        print(f"M3 回归测试失败，共 {len(failures)} 项不一致：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("M3 回归测试全部通过 ✅（统计/明细DF/更新表ID/颜色/说明md/输入校验）")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        traceback.print_exc()
        sys.exit(2)
