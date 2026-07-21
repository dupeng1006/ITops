# -*- coding: utf-8 -*-
"""
M2 黄金样本回归测试（纯脚本，无 pytest 依赖）

比对项：
    1. 统计 dict：与 expected/expected_stats.json 逐项相等（products + 合计）；
    2. 逐行判定：与 expected/expected_rows.json 逐行比对
       （证券代码/证券名称/差异状态/匹配方式/科目类型/备注）；
    3. 报告列结构：openpyxl 读实际报告表头 == REPORT_COLUMNS（规格 10 列 + 备注）；
    4. 三类颜色抽样：一致行 绿C6EFCE / 差异行 红FFC7CE / 单边行 橙FFC000；
    5. 底部汇总行：汇总/总记录/一致/差异/单边 数值与统计一致；
    6. 1501 取价金丝雀：110103 估值表价格必须为单位成本 100.0000
       （若误取市价 102.5000 则此断言失败，防止取价字段被改错）；
    7. 备注专项：1501 差异行备注含"摊余成本"；888880 单边行备注含"新标准券"；
    8. 输入校验单测：系统端缺列/估值表缺列/缺取价字段列/空规则运行，
       均抛 ValueError 且中文提示指明问题。

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests/golden/m2/test_golden_m2.py
退出码：全部通过 0；任一不一致打印差异明细并以非零码退出。

作者：技术部
版本：1.0.0
日期：2026-07-20
"""

import json
import sys
import tempfile
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# 将 server/ 加入模块搜索路径（本文件位于 server/tests/golden/m2/）
SERVER_ROOT = Path(__file__).resolve().parents[3]
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

from app.engines.m2_valuation_price import (  # noqa: E402
    FILL_DIFF,
    FILL_SAME,
    FILL_SINGLE,
    REPORT_COLUMNS,
    M2ValuationPriceEngine,
    SubjectPriceRuleConfig,
)

M2_DIR = Path(__file__).resolve().parent
SAMPLE_DIR = M2_DIR / "samples"
EXPECTED_DIR = M2_DIR / "expected"
ACTUAL_DIR = M2_DIR / "actual"

SYS_6301 = SAMPLE_DIR / "新综合信息查询_基金证券-6301.xlsx"
VAL_6301 = SAMPLE_DIR / "证券投资基金估值表_6301-20260720.xlsx"
SYS_6302 = SAMPLE_DIR / "新综合信息查询_基金证券-6302.xlsx"
VAL_6302 = SAMPLE_DIR / "证券投资基金估值表_6302-20260720.xlsx"

EXPECTED_STATS = EXPECTED_DIR / "expected_stats.json"
EXPECTED_ROWS = EXPECTED_DIR / "expected_rows.json"

# 黄金口径（与 build_expected_m2.py 一致；变更须先评审并重建基线）
GOLDEN_RULES = [
    SubjectPriceRuleConfig("1101", "市价", "交易性金融资产", "", 1),
    SubjectPriceRuleConfig("1501", "单位成本", "债权投资",
                           "摊余成本与市场估值口径差异，属正常", 2),
]
GOLDEN_FUZZY_SIM = 0.5
GOLDEN_PRICE_TOL = 0.0001

ROW_KEYS = ["证券代码", "证券名称", "差异状态", "匹配方式", "科目类型", "备注"]

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def cell_rgb(ws, row: int, col: int) -> str:
    return ws.cell(row=row, column=col).fill.start_color.rgb or ""


def make_engine() -> M2ValuationPriceEngine:
    return M2ValuationPriceEngine(
        subject_rules=list(GOLDEN_RULES),
        fuzzy_sim=GOLDEN_FUZZY_SIM,
        price_tol=GOLDEN_PRICE_TOL,
    )


# =============================================================================
# 检查 1：统计 dict
# =============================================================================

def check_stats(actual_stats: dict) -> None:
    with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
        expected_stats = json.load(f)
    check("统计 dict 逐项相等", actual_stats == expected_stats,
          f"预期{expected_stats} != 实际{actual_stats}")


# =============================================================================
# 检查 2：逐行判定（六键比对）
# =============================================================================

def check_rows(results: list) -> None:
    with open(EXPECTED_ROWS, "r", encoding="utf-8") as f:
        expected_rows = json.load(f)
    actual_by_product = {r["product"]: r["report_df"] for r in results}
    check("产品集合一致", set(actual_by_product) == set(expected_rows),
          f"预期{sorted(expected_rows)} != 实际{sorted(actual_by_product)}")
    for product, exp_rows in expected_rows.items():
        df = actual_by_product.get(product)
        if df is None:
            continue
        check(f"{product} 报告行数={len(exp_rows)}", len(df) == len(exp_rows),
              f"预期{len(exp_rows)} != 实际{len(df)}")
        mismatch = []
        for idx, exp in enumerate(exp_rows):
            if idx >= len(df):
                break
            row = df.iloc[idx]
            for key in ROW_KEYS:
                actual_val = row[key]
                actual_val = "" if actual_val is None or (
                    isinstance(actual_val, float) and pd.isna(actual_val)) else str(actual_val)
                if actual_val != str(exp[key]):
                    mismatch.append(
                        f"{product} 行{idx}[{exp['证券代码']}] 列[{key}]: "
                        f"预期{exp[key]!r} != 实际{actual_val!r}")
        check(f"{product} 逐行六键比对一致", not mismatch, "\n    ".join(mismatch[:20]))


# =============================================================================
# 检查 3：报告列结构 + 检查 4：颜色抽样 + 检查 5：底部汇总行
# =============================================================================

def check_report_excel(results: list) -> None:
    with open(EXPECTED_ROWS, "r", encoding="utf-8") as f:
        expected_rows = json.load(f)
    status_color = {"一致": FILL_SAME, "差异": FILL_DIFF,
                    "系统有估值表无": FILL_SINGLE, "估值表有系统无": FILL_SINGLE}
    for r in results:
        product = r["product"]
        path = Path(r["output_file"])
        wb = load_workbook(path)
        ws = wb.active
        # 列结构
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(REPORT_COLUMNS))]
        check(f"{product} 报告表头==REPORT_COLUMNS", headers == REPORT_COLUMNS,
              f"实际{headers}")
        # 颜色逐行（依据预期逐行差异状态判定色值）
        exp_rows = expected_rows[product]
        mismatch = []
        for idx, exp in enumerate(exp_rows):
            want = status_color[exp["差异状态"]]
            got = cell_rgb(ws, idx + 2, 1)
            if not got.endswith(want):
                mismatch.append(f"行{idx + 2}[{exp['证券代码']}]: 预期色{want} 实际{got}")
        check(f"{product} 全部行颜色与差异状态匹配", not mismatch,
              "\n    ".join(mismatch[:20]))
        # 底部汇总行（数据行数 + 3 行处：空一行后写入）
        srow = len(exp_rows) + 3
        stats = r["stats"]
        summary_vals = [ws.cell(row=srow, column=c).value for c in range(1, 6)]
        want_vals = ["汇总", f"总记录 {stats['总记录']}", f"一致 {stats['一致']}",
                     f"差异 {stats['差异']}", f"单边 {stats['单边']}"]
        check(f"{product} 底部汇总行值正确", summary_vals == want_vals,
              f"预期{want_vals} != 实际{summary_vals}")


# =============================================================================
# 检查 6：1501 取价金丝雀（单位成本 100.0000，非市价 102.5000）
# =============================================================================

def check_price_canary(results: list) -> None:
    df_6301 = {r["product"]: r["report_df"] for r in results}["6301"]
    row = df_6301[df_6301["证券代码"] == "110103"]
    check("金丝雀行 110103 存在", len(row) == 1, f"实际{len(row)}行")
    if len(row) != 1:
        return
    val_price = float(row.iloc[0]["估值表价格"])
    check("1501 取价=单位成本 100.0000", abs(val_price - 100.0) < 1e-9,
          f"实际估值表价格={val_price}（若为 102.5 说明误取市价列）")
    diff = float(row.iloc[0]["价格差异"])
    check("1501 价格差异=2.0", abs(diff - 2.0) < 1e-9, f"实际差异={diff}")


# =============================================================================
# 检查 7：备注专项
# =============================================================================

def check_notes(results: list) -> None:
    by_product = {r["product"]: r["report_df"] for r in results}
    df_6301 = by_product["6301"]
    row_1501 = df_6301[df_6301["证券代码"] == "110103"].iloc[0]
    check("1501 差异行备注含[摊余成本]", "摊余成本" in str(row_1501["备注"]),
          f"实际备注={row_1501['备注']!r}")
    row_888 = df_6301[df_6301["证券代码"] == "888880"].iloc[0]
    check("888880 单边行备注含[新标准券]", "新标准券" in str(row_888["备注"]),
          f"实际备注={row_888['备注']!r}")
    df_6302 = by_product["6302"]
    row_220202 = df_6302[df_6302["证券代码"] == "220202"].iloc[0]
    check("6302 的 1501 差异行备注含[摊余成本]",
          "摊余成本" in str(row_220202["备注"]), f"实际备注={row_220202['备注']!r}")


# =============================================================================
# 检查 8：输入校验单测（中文报错）
# =============================================================================

def check_validation() -> None:
    engine = make_engine()
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        # 系统端缺 持仓 列
        bad_sys = tmp / "缺列系统端.xlsx"
        pd.DataFrame({"证券代码": ["110101"], "证券名称": ["测试"], "估值价格": [100.0]}
                     ).to_excel(bad_sys, index=False)
        try:
            engine.load_system_file(bad_sys)
            check("系统端缺列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("系统端缺列抛 ValueError", True)
            check("报错指明缺[持仓]列", "持仓" in str(e), str(e))

        # 估值表缺 数量 列（skiprows=3，列头在第 4 行）
        bad_val = tmp / "缺列估值表.xlsx"
        pd.DataFrame({"科目代码": ["110101000000110101"], "科目名称": ["测试"],
                      "市价": [100.0]}).to_excel(bad_val, index=False, startrow=3)
        try:
            engine.load_valuation_file(bad_val)
            check("估值表缺列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("估值表缺列抛 ValueError", True)
            check("报错指明缺[数量]列", "数量" in str(e), str(e))

        # 估值表缺规则取价字段列（有数量/市价但无单位成本）
        no_field = tmp / "缺取价字段估值表.xlsx"
        pd.DataFrame({"科目代码": ["110101000000110101"], "科目名称": ["测试"],
                      "数量": [1000], "市价": [100.0]}
                     ).to_excel(no_field, index=False, startrow=3)
        try:
            engine.load_valuation_file(no_field, price_fields=["单位成本"])
            check("缺取价字段列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("缺取价字段列抛 ValueError", True)
            check("报错指明缺[单位成本]列", "单位成本" in str(e), str(e))

        # 空规则运行
        empty_engine = M2ValuationPriceEngine(subject_rules=[])
        try:
            empty_engine.run(
                jobs=[{"product": "6301", "system_path": str(SYS_6301),
                       "valuation_path": str(VAL_6301)}],
                output_dir=str(tmp / "out"))
            check("空规则运行抛 ValueError", False, "未抛异常")
        except ValueError as e:
            check("空规则运行抛 ValueError", True)
            check("报错提示配置科目取价规则", "科目取价规则" in str(e), str(e))


# =============================================================================
# 主流程
# =============================================================================

def main() -> int:
    print("=" * 70)
    print("M2 黄金样本回归测试（基金估值价格核对）")
    print("=" * 70)

    for path in (SYS_6301, VAL_6301, SYS_6302, VAL_6302,
                 EXPECTED_STATS, EXPECTED_ROWS):
        if not path.exists():
            print(f"缺失文件: {path}，请先运行 make_samples_m2.py 与 build_expected_m2.py")
            return 2

    ACTUAL_DIR.mkdir(parents=True, exist_ok=True)
    engine = make_engine()
    result = engine.run(
        jobs=[
            {"product": "6301", "system_path": str(SYS_6301),
             "valuation_path": str(VAL_6301)},
            {"product": "6302", "system_path": str(SYS_6302),
             "valuation_path": str(VAL_6302)},
        ],
        output_dir=str(ACTUAL_DIR),
    )
    print("-" * 70)
    print("【检查 1】统计 dict")
    check_stats(result["stats"])
    print("【检查 2】逐行判定（六键比对）")
    check_rows(result["results"])
    print("【检查 3-5】报告列结构 / 颜色抽样 / 底部汇总行")
    check_report_excel(result["results"])
    print("【检查 6】1501 取价金丝雀")
    check_price_canary(result["results"])
    print("【检查 7】备注专项")
    check_notes(result["results"])
    print("【检查 8】输入校验单测")
    check_validation()

    print("=" * 70)
    if failures:
        print(f"M2 回归测试失败，共 {len(failures)} 项不一致：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("M2 回归测试全部通过 ✅（统计/逐行/列结构/颜色/汇总行/金丝雀/备注/输入校验）")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        traceback.print_exc()
        sys.exit(2)
