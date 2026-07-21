# -*- coding: utf-8 -*-
"""
M1 黄金样本回归测试（纯脚本，无 pytest 依赖）

比对项：
    1. 统计 dict：与 expected/expected_stats.json 逐项相等；
    2. 结果 DataFrame：与 expected/expected_result_df.pkl 逐行逐列值相等
       （浮点容差 1e-6，NaN 与 NaN 视为相等）；
    3. 输出 Excel 填充色：与 expected/核对结果_预期.xlsx 逐格比对
       fill.start_color.rgb（表头至末行 × 15 列）。

附加单测：
    4. 列数不足报错：基金资产表 <28 列、净值查询表 <9 列分别抛 ValueError，
       中文提示包含"选反"引导语；
    5. 文件选反告警：两表列数特征疑似选反时记录告警日志（含"选反"），
       且随后抛出列数不足 ValueError。

运行：
    python server/tests/golden/test_golden.py
退出码：全部通过 0；任一不一致打印差异明细并以非零码退出。

作者：技术部
版本：1.0.0
日期：2026-07-17
"""

import io
import json
import logging
import pickle
import sys
import tempfile
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# 将 server/ 加入模块搜索路径（本文件位于 server/tests/golden/）
SERVER_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = SERVER_ROOT.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

from app.engines.m1_fund_netvalue import M1FundNetvalueEngine  # noqa: E402

GOLDEN_DIR = Path(__file__).resolve().parent
EXPECTED_DIR = GOLDEN_DIR / "expected"
ACTUAL_DIR = GOLDEN_DIR / "actual"

FUND_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "基金资产表_样本.xlsx"
NETVALUE_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "净值查询表_样本.xlsx"

EXPECTED_EXCEL = EXPECTED_DIR / "核对结果_预期.xlsx"
EXPECTED_STATS = EXPECTED_DIR / "expected_stats.json"
EXPECTED_DF = EXPECTED_DIR / "expected_result_df.pkl"
ACTUAL_EXCEL = ACTUAL_DIR / "核对结果_实际.xlsx"

FLOAT_TOL = 1e-6  # 金额等浮点比对容差

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    """记录单项检查结果"""
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def values_equal(expected, actual, tol: float = FLOAT_TOL) -> bool:
    """单元格值比对：NaN==NaN，数值容差比对，其余严格相等"""
    exp_na = pd.isna(expected) if not isinstance(expected, (list, dict)) else False
    act_na = pd.isna(actual) if not isinstance(actual, (list, dict)) else False
    if exp_na or act_na:
        return bool(exp_na and act_na)
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(expected) - float(actual)) <= tol
    return expected == actual


# =============================================================================
# 检查 1：统计 dict
# =============================================================================

def check_stats(actual_stats: dict) -> None:
    with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
        expected_stats = json.load(f)

    if set(expected_stats) != set(actual_stats):
        check("统计 dict 键集合一致", False,
              f"预期键{sorted(expected_stats)} != 实际键{sorted(actual_stats)}")
        return
    check("统计 dict 键集合一致", True)

    for key in expected_stats:
        check(f"统计项[{key}] 预期={expected_stats[key]} 实际={actual_stats.get(key)}",
              expected_stats[key] == actual_stats.get(key),
              f"预期 {expected_stats[key]!r} != 实际 {actual_stats.get(key)!r}")


# =============================================================================
# 检查 2：结果 DataFrame 逐行逐列
# =============================================================================

def check_dataframe(actual_df: pd.DataFrame) -> None:
    with open(EXPECTED_DF, "rb") as f:
        expected_df = pickle.load(f)

    check("DataFrame 形状一致", expected_df.shape == actual_df.shape,
          f"预期{expected_df.shape} != 实际{actual_df.shape}")
    if expected_df.shape != actual_df.shape:
        return
    check("DataFrame 列名一致", list(expected_df.columns) == list(actual_df.columns),
          f"预期{list(expected_df.columns)} != 实际{list(actual_df.columns)}")

    mismatch = []
    for col in expected_df.columns:
        for idx in range(len(expected_df)):
            exp_val = expected_df.iloc[idx][col]
            act_val = actual_df.iloc[idx][col]
            if not values_equal(exp_val, act_val):
                mismatch.append(
                    f"行{idx} 列[{col}]: 预期{exp_val!r} != 实际{act_val!r}"
                )
    check(f"DataFrame 逐行逐列值相等（容差{FLOAT_TOL:g}）", not mismatch,
          "\n    ".join(mismatch[:20]) + (f"\n    ...共{len(mismatch)}处" if len(mismatch) > 20 else ""))


# =============================================================================
# 检查 3：输出 Excel 填充色逐格比对
# =============================================================================

def check_excel_fills() -> None:
    wb_exp = load_workbook(EXPECTED_EXCEL)
    wb_act = load_workbook(ACTUAL_EXCEL)
    ws_exp, ws_act = wb_exp.active, wb_act.active

    check("Excel Sheet 名一致", ws_exp.title == ws_act.title,
          f"预期{ws_exp.title!r} != 实际{ws_act.title!r}")
    check("Excel 尺寸一致", ws_exp.dimensions == ws_act.dimensions,
          f"预期{ws_exp.dimensions} != 实际{ws_act.dimensions}")
    if ws_exp.dimensions != ws_act.dimensions:
        return

    mismatch = []
    for r in range(1, ws_exp.max_row + 1):
        for c in range(1, 16):
            exp_rgb = ws_exp.cell(row=r, column=c).fill.start_color.rgb
            act_rgb = ws_act.cell(row=r, column=c).fill.start_color.rgb
            if exp_rgb != act_rgb:
                mismatch.append(
                    f"({r},{c}) {ws_exp.cell(row=r, column=2).value}: "
                    f"预期{exp_rgb} != 实际{act_rgb}"
                )
    check("Excel 单元格填充色逐格一致", not mismatch,
          "\n    ".join(mismatch[:20]) + (f"\n    ...共{len(mismatch)}处" if len(mismatch) > 20 else ""))


# =============================================================================
# 检查 4/5：输入校验单测
# =============================================================================

def _make_logger_with_capture():
    """构造带内存捕获的日志记录器，返回 (logger, buffer)"""
    buffer = io.StringIO()
    logger = logging.getLogger("o32ops.m1.unittest")
    logger.handlers = []
    handler = logging.StreamHandler(buffer)
    handler.setFormatter(logging.Formatter("%(levelname)s - %(message)s"))
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    logger.propagate = False
    return logger, buffer


def _write_dummy_excel(path: Path, ncols: int) -> None:
    """生成仅含表头的占位 Excel（列数由 ncols 决定）"""
    pd.DataFrame(columns=[f"列{i}" for i in range(ncols)]).to_excel(path, index=False)


def check_validation() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        fund_ok = tmp / "fund_ok.xlsx"     # 28 列
        fund_short = tmp / "fund_short.xlsx"  # 20 列（<28 但不触发选反特征）
        net_ok = tmp / "net_ok.xlsx"       # 10 列（>=9）
        net_short = tmp / "net_short.xlsx"    # 8 列（<9）
        net_like = tmp / "net_like.xlsx"   # 9 列（伪装成基金表）
        fund_like = tmp / "fund_like.xlsx"  # 28 列（伪装成净值表）
        out = tmp / "out.xlsx"

        _write_dummy_excel(fund_ok, 28)
        _write_dummy_excel(fund_short, 20)
        _write_dummy_excel(net_ok, 10)
        _write_dummy_excel(net_short, 8)
        _write_dummy_excel(net_like, 9)
        _write_dummy_excel(fund_like, 28)

        # 4a. 基金资产表列数不足
        logger, _ = _make_logger_with_capture()
        engine = M1FundNetvalueEngine(logger=logger)
        try:
            engine.run(fund_short, net_ok, out)
            check("基金资产表<28列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            msg = str(e)
            check("基金资产表<28列抛 ValueError", True)
            check("报错提示含'列数不足'", "列数不足" in msg, msg)
            check("报错提示含'选反'引导语", "选反" in msg, msg)
        except Exception as e:  # noqa: BLE001
            check("基金资产表<28列抛 ValueError", False, f"异常类型错误: {type(e).__name__}: {e}")

        # 4b. 净值查询表列数不足
        logger, _ = _make_logger_with_capture()
        engine = M1FundNetvalueEngine(logger=logger)
        try:
            engine.run(fund_ok, net_short, out)
            check("净值查询表<9列抛 ValueError", False, "未抛异常")
        except ValueError as e:
            msg = str(e)
            check("净值查询表<9列抛 ValueError", True)
            check("净值表报错提示含'选反'引导语", "选反" in msg, msg)
        except Exception as e:  # noqa: BLE001
            check("净值查询表<9列抛 ValueError", False, f"异常类型错误: {type(e).__name__}: {e}")

        # 5. 文件选反：告警日志 + ValueError
        logger, buffer = _make_logger_with_capture()
        engine = M1FundNetvalueEngine(logger=logger)
        try:
            engine.run(net_like, fund_like, out)  # 两文件位置故意选反
            check("文件选反场景抛 ValueError", False, "未抛异常")
        except ValueError as e:
            logs = buffer.getvalue()
            check("文件选反场景抛 ValueError", True)
            check("选反场景记录告警日志（含'选反'）", "选反" in logs and "WARNING" in logs,
                  f"日志内容: {logs!r}")
            check("选反场景报错提示含'选反'引导语", "选反" in str(e), str(e))
        except Exception as e:  # noqa: BLE001
            check("文件选反场景抛 ValueError", False, f"异常类型错误: {type(e).__name__}: {e}")


# =============================================================================
# 主流程
# =============================================================================

def main() -> int:
    print("=" * 70)
    print("M1 黄金样本回归测试")
    print("=" * 70)

    for path in (FUND_SAMPLE, NETVALUE_SAMPLE, EXPECTED_EXCEL, EXPECTED_STATS, EXPECTED_DF):
        if not path.exists():
            print(f"缺失文件: {path}，请先运行 make_samples.py 与 build_expected.py")
            return 2

    # 重新执行引擎
    ACTUAL_DIR.mkdir(parents=True, exist_ok=True)
    engine = M1FundNetvalueEngine()
    result = engine.run(
        fund_path=str(FUND_SAMPLE),
        netvalue_path=str(NETVALUE_SAMPLE),
        output_path=str(ACTUAL_EXCEL),
    )
    print("-" * 70)
    print("【检查 1】统计 dict 逐项比对")
    check_stats(result["stats"])
    print("-" * 70)
    print("【检查 2】结果 DataFrame 逐行逐列比对")
    check_dataframe(result["result_df"])
    print("-" * 70)
    print("【检查 3】输出 Excel 填充色逐格比对")
    check_excel_fills()
    print("-" * 70)
    print("【检查 4/5】输入校验单测（列数不足报错 / 文件选反告警）")
    check_validation()

    print("=" * 70)
    if failures:
        print(f"回归测试失败，共 {len(failures)} 项不一致：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("回归测试全部通过 ✅（统计 dict / DataFrame 逐值 / Excel 填充色 / 输入校验）")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        traceback.print_exc()
        sys.exit(2)
