# -*- coding: utf-8 -*-
"""
M3 API 冒烟测试（纯脚本，fastapi.testclient + 临时目录数据/归档）

覆盖链路：
    1. admin 登录改密 → 创建 operator / viewer
    2. operator 上传黄金样本（基金属性表 xlsx + 交易成员表 GBK CSV）建 M3 任务
       → 轮询至 success → 统计摘要与 M3 黄金基线一致
    3. 任务详情返回三件套清单
    4. 下载三件套（?file=updated|detail|note）：颜色抽样（蓝/绿/红）、
       更新表银行间ID 值、明细内容、说明 md 关键内容
    5. 默认下载（不带 file 参数）= 更新表
    6. 缺列文件（基金属性表缺 银行间ID / 交易成员表缺 交易成员ID）
       → 400 中文提示含缺失列名
    7. viewer 建 M3 任务 403、下载 403、可查历史 200

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests/api/test_api_m3.py
退出码：全部通过 0；任一失败非零。

作者：技术部
版本：1.0.0
日期：2026-07-17
"""

import json
import os
import sys
import tempfile
import time
from pathlib import Path

# =============================================================================
# 环境准备：导入 app 之前注入测试配置（临时数据/归档目录、测试密钥）
# =============================================================================

SERVER_ROOT = Path(__file__).resolve().parents[2]
TMP_ROOT = Path(tempfile.mkdtemp(prefix="o32ops_m3_smoke_"))

os.environ["O32OPS_DATA_DIR"] = str(TMP_ROOT / "data")
os.environ["O32OPS_ARCHIVE_DIR"] = str(TMP_ROOT / "archive")
os.environ["O32OPS_DB_PATH"] = str(TMP_ROOT / "data" / "o32ops.db")
os.environ["O32OPS_SECRET_KEY"] = "m3-smoke-test-secret-key-do-not-use-in-prod"

if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

import pandas as pd  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402

M3_GOLDEN_DIR = SERVER_ROOT / "tests" / "golden" / "m3"
FUND_SAMPLE = M3_GOLDEN_DIR / "samples" / "基金属性表_样本.xlsx"
MEMBER_SAMPLE = M3_GOLDEN_DIR / "samples" / "交易成员基本信息表_样本.csv"
EXPECTED_STATS = M3_GOLDEN_DIR / "expected" / "expected_stats.json"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME = "text/csv"

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login_and_change(client: TestClient, username: str, old_pw: str, new_pw: str) -> str:
    resp = client.post("/api/auth/login", json={"username": username, "password": old_pw})
    check(f"登录 {username} 返回200", resp.status_code == 200, resp.text)
    token = resp.json()["access_token"]
    if resp.json().get("must_change_password"):
        resp2 = client.post("/api/auth/change-password",
                            json={"old_password": old_pw, "new_password": new_pw},
                            headers=auth(token))
        check(f"{username} 修改密码返回200", resp2.status_code == 200, resp2.text)
    return token


def create_m3_job(client: TestClient, token: str) -> str:
    resp = client.post(
        "/api/recon/m3/jobs",
        files={
            "fund_file": (FUND_SAMPLE.name, FUND_SAMPLE.read_bytes(), XLSX_MIME),
            "member_file": (MEMBER_SAMPLE.name, MEMBER_SAMPLE.read_bytes(), CSV_MIME),
        },
        headers=auth(token),
    )
    check("创建 M3 任务返回200", resp.status_code == 200, resp.text)
    return resp.json()["job_id"]


def poll_job(client: TestClient, token: str, job_id: str, timeout: float = 90.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/recon/jobs/{job_id}", headers=auth(token))
        job = resp.json()
        if job["status"] in ("success", "failed"):
            return job
        time.sleep(0.3)
    raise TimeoutError(f"任务 {job_id} 轮询超时")


def main() -> int:
    print("=" * 70)
    print("M3 API 冒烟测试（基金属性表银行间ID匹配）")
    print(f"临时目录: {TMP_ROOT}")
    print("=" * 70)

    with TestClient(app) as client:
        # 1. 用户准备
        print("--- 1. 用户准备 ---")
        admin_token = login_and_change(client, "admin", "Admin@123", "Admin@2026New")
        resp = client.post("/api/admin/users",
                           json={"username": "op03", "password": "Op@123456", "role": "operator"},
                           headers=auth(admin_token))
        check("创建 operator 返回200", resp.status_code == 200, resp.text)
        resp = client.post("/api/admin/users",
                           json={"username": "vw03", "password": "Vw@123456", "role": "viewer"},
                           headers=auth(admin_token))
        check("创建 viewer 返回200", resp.status_code == 200, resp.text)
        op_token = login_and_change(client, "op03", "Op@123456", "Op@654321New")
        vw_token = login_and_change(client, "vw03", "Vw@123456", "Vw@654321New")

        # 2. M3 任务全流程
        print("--- 2. M3 任务全流程 ---")
        job_id = create_m3_job(client, op_token)
        final = poll_job(client, op_token, job_id)
        check("M3 任务执行成功", final["status"] == "success",
              f"status={final['status']} error={final.get('error')}")

        with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
            expected_stats = json.load(f)
        check("统计摘要与 M3 黄金基线一致", final["stats"] == expected_stats,
              f"预期{expected_stats} 实际{final['stats']}")

        # 3. 任务详情三件套清单
        print("--- 3. 任务详情 ---")
        result_files = final.get("result_files") or []
        for name in ("基金属性_精确匹配更新.xlsx", "精确匹配结果明细.xlsx", "精确匹配说明.md"):
            check(f"结果文件清单含[{name}]", name in result_files, str(result_files))
        check("详情日志含中文步骤文案",
              any("精确匹配完成" in line for line in final["log_tail"]),
              str(final["log_tail"][-3:]))

        # 4. 下载三件套
        print("--- 4. 下载三件套与内容抽样 ---")
        # 4.1 更新表
        resp = client.get(f"/api/recon/jobs/{job_id}/download",
                          params={"file": "updated"}, headers=auth(op_token))
        check("下载 updated 返回200", resp.status_code == 200, resp.text)
        path = TMP_ROOT / "updated.xlsx"
        path.write_bytes(resp.content)
        ws = load_workbook(path).active
        rgb = lambda r: ws.cell(row=r, column=1).fill.start_color.rgb or ""
        check("更新表 第2行蓝色(无变化)", rgb(2).endswith("BDD7EE"), rgb(2))
        check("更新表 第6行绿色(有变动)", rgb(6).endswith("C6EFCE"), rgb(6))
        check("更新表 第11行红色(未匹配)", rgb(11).endswith("FFC7CE"), rgb(11))
        check("更新表 银行间ID 第6行=MB1005",
              ws.cell(row=6, column=4).value == "MB1005", str(ws.cell(row=6, column=4).value))
        check("更新表 银行间ID 第2行=1001(数值陷阱归一)",
              str(ws.cell(row=2, column=4).value) == "1001", str(ws.cell(row=2, column=4).value))

        # 4.2 明细表
        resp = client.get(f"/api/recon/jobs/{job_id}/download",
                          params={"file": "detail"}, headers=auth(op_token))
        check("下载 detail 返回200", resp.status_code == 200, resp.text)
        path = TMP_ROOT / "detail.xlsx"
        path.write_bytes(resp.content)
        ws = load_workbook(path).active
        check("明细表 第2行匹配类型=精确匹配-无变化",
              ws.cell(row=2, column=4).value == "精确匹配-无变化", str(ws.cell(row=2, column=4).value))
        check("明细表 第11行匹配类型=未匹配",
              ws.cell(row=11, column=4).value == "未匹配", str(ws.cell(row=11, column=4).value))

        # 4.3 说明 md
        resp = client.get(f"/api/recon/jobs/{job_id}/download",
                          params={"file": "note"}, headers=auth(op_token))
        check("下载 note 返回200", resp.status_code == 200, resp.text)
        note_text = resp.content.decode("utf-8")
        for keyword in ("总记录数: 15", "精确匹配-有变动: 5", "未匹配: 6",
                        "内部简称（TR_ 前缀）", "命名后缀不一致", "未注册产品"):
            check(f"说明 md 含[{keyword}]", keyword in note_text, keyword)

        # 4.4 默认下载 = 更新表
        resp = client.get(f"/api/recon/jobs/{job_id}/download", headers=auth(op_token))
        check("默认下载返回200", resp.status_code == 200, resp.text)
        check("默认下载为 xlsx(PK头)", resp.content[:2] == b"PK", f"长度={len(resp.content)}")

        # 4.5 非法 file 参数
        resp = client.get(f"/api/recon/jobs/{job_id}/download",
                          params={"file": "bad"}, headers=auth(op_token))
        check("非法 file 参数返回400", resp.status_code == 400, resp.text)

        # 5. viewer 权限边界
        print("--- 5. viewer 权限边界 ---")
        resp = client.post(
            "/api/recon/m3/jobs",
            files={"fund_file": (FUND_SAMPLE.name, FUND_SAMPLE.read_bytes(), XLSX_MIME),
                   "member_file": (MEMBER_SAMPLE.name, MEMBER_SAMPLE.read_bytes(), CSV_MIME)},
            headers=auth(vw_token),
        )
        check("viewer 建 M3 任务被拒(403)", resp.status_code == 403, resp.text)
        resp = client.get(f"/api/recon/jobs/{job_id}/download", headers=auth(vw_token))
        check("viewer 下载被拒(403)", resp.status_code == 403, resp.text)
        resp = client.get("/api/recon/jobs", params={"module": "M3"}, headers=auth(vw_token))
        check("viewer 可查 M3 历史(200)",
              resp.status_code == 200 and resp.json()["total"] == 1, resp.text)

        # 6. 缺列文件 400
        print("--- 6. 缺列文件 400 ---")
        bad_fund = TMP_ROOT / "缺列基金属性表.xlsx"
        pd.DataFrame({"基金代码": ["F001"], "基金全称": ["测试产品"]}).to_excel(bad_fund, index=False)
        resp = client.post(
            "/api/recon/m3/jobs",
            files={"fund_file": (bad_fund.name, bad_fund.read_bytes(), XLSX_MIME),
                   "member_file": (MEMBER_SAMPLE.name, MEMBER_SAMPLE.read_bytes(), CSV_MIME)},
            headers=auth(op_token),
        )
        check("基金属性表缺列返回400", resp.status_code == 400, resp.text)
        check("400 提示含缺失列名[银行间ID]", "银行间ID" in resp.text, resp.text)

        bad_member = TMP_ROOT / "缺列交易成员表.csv"
        pd.DataFrame({"交易成员全称": ["测试成员"], "机构代码": ["ORG001"]}).to_csv(
            bad_member, index=False, encoding="gbk")
        resp = client.post(
            "/api/recon/m3/jobs",
            files={"fund_file": (FUND_SAMPLE.name, FUND_SAMPLE.read_bytes(), XLSX_MIME),
                   "member_file": (bad_member.name, bad_member.read_bytes(), CSV_MIME)},
            headers=auth(op_token),
        )
        check("交易成员表缺列返回400", resp.status_code == 400, resp.text)
        check("400 提示含缺失列名[交易成员ID]", "交易成员ID" in resp.text, resp.text)

    print("=" * 70)
    if failures:
        print(f"M3 冒烟测试失败，共 {len(failures)} 项：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("M3 API 冒烟测试全部通过 ✅")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        sys.exit(2)
