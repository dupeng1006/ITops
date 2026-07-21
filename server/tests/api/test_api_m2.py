# -*- coding: utf-8 -*-
"""
M2 API 冒烟测试（纯脚本，fastapi.testclient + 临时目录数据/归档）

覆盖链路：
    1. admin 登录改密 → 创建 operator / viewer
    2. 科目取价规则种子核验（GET 列表含 1101→市价 / 1501→单位成本）
    3. operator 上传黄金样本 4 文件（系统端×2 + 估值表×2，多文件字段）建 M2 任务
       → 轮询至 success → 统计摘要与 M2 黄金基线一致（6301: 7/2/2/3，6302: 2/1/1/0）
    4. 按产品下载（?product=6301 / 6302）：表头结构、三色抽样、底部汇总行、
       1501 取价金丝雀（估值表价格=100.0）、备注关键词；缺省下载=首个产品报告；
       ?product=9999 → 404 中文提示
    5. 配对异常 400：落单（2 系统端 + 1 估值表，指明落单文件）/ 文件名无产品标识 /
       系统端缺列（校验失败清理，提示缺列名）
    6. viewer 权限边界：建任务 403 / 查规则 403 / 查历史 200
    7. 规则 CRUD 权限与唯一性：operator 新增 403 / 重复前缀 400 / PUT 修改留痕
    8. 热生效专项：admin 新增 1102→市价 规则 → 重跑任务 6301 变 8/2/2/4
       且报告含「1102-其他投资（取市价）」行 → DELETE 规则 → 重跑回基线
    9. 审计日志落库核验（sys_subject_rule_create/update/delete、upload_create_job）

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests/api/test_api_m2.py
退出码：全部通过 0；任一失败非零。

作者：技术部
版本：1.0.0
日期：2026-07-20
"""

import json
import os
import sqlite3
import sys
import tempfile
import time
from pathlib import Path

# =============================================================================
# 环境准备：导入 app 之前注入测试配置（临时数据/归档目录、测试密钥）
# =============================================================================

SERVER_ROOT = Path(__file__).resolve().parents[2]
TMP_ROOT = Path(tempfile.mkdtemp(prefix="o32ops_m2_smoke_"))

os.environ["O32OPS_DATA_DIR"] = str(TMP_ROOT / "data")
os.environ["O32OPS_ARCHIVE_DIR"] = str(TMP_ROOT / "archive")
os.environ["O32OPS_DB_PATH"] = str(TMP_ROOT / "data" / "o32ops.db")
os.environ["O32OPS_SECRET_KEY"] = "m2-smoke-test-secret-key-do-not-use-in-prod"

if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

import pandas as pd  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.engines.m2_valuation_price import REPORT_COLUMNS  # noqa: E402
from app.main import app  # noqa: E402

M2_GOLDEN_DIR = SERVER_ROOT / "tests" / "golden" / "m2"
SAMPLE_DIR = M2_GOLDEN_DIR / "samples"
SYS_6301 = SAMPLE_DIR / "新综合信息查询_基金证券-6301.xlsx"
SYS_6302 = SAMPLE_DIR / "新综合信息查询_基金证券-6302.xlsx"
VAL_6301 = SAMPLE_DIR / "证券投资基金估值表_6301-20260720.xlsx"
VAL_6302 = SAMPLE_DIR / "证券投资基金估值表_6302-20260720.xlsx"
EXPECTED_STATS = M2_GOLDEN_DIR / "expected" / "expected_stats.json"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login_and_change(client: TestClient, username: str, old_pw: str, new_pw: str) -> str:
    resp = client.post("/api/auth/login", json={"username": username, "password": old_pw})
    check(f"登录 {username} 返回200", resp.status_code == 200, resp.text)
    token = resp.json()["access_token"]
    if resp.json().get("must_change_password"):
        resp2 = client.post("/api/auth/change-password",
                            json={"old_password": old_pw, "new_password": new_pw},
                            headers=auth(token))
        check(f"{username} 修改密码返回200", resp2.status_code == 200, resp2.text)
    return token


def m2_files(sys_paths, val_paths):
    """构造多文件 multipart（system_files / valuation_files 同名字段列表）"""
    files = []
    for p in sys_paths:
        files.append(("system_files", (p.name, p.read_bytes(), XLSX_MIME)))
    for p in val_paths:
        files.append(("valuation_files", (p.name, p.read_bytes(), XLSX_MIME)))
    return files


def create_m2_job(client: TestClient, token: str) -> str:
    resp = client.post("/api/recon/m2/jobs",
                       files=m2_files([SYS_6301, SYS_6302], [VAL_6301, VAL_6302]),
                       headers=auth(token))
    check("创建 M2 任务返回200", resp.status_code == 200, resp.text)
    return resp.json()["job_id"]


def poll_job(client: TestClient, token: str, job_id: str, timeout: float = 90.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/recon/jobs/{job_id}", headers=auth(token))
        job = resp.json()
        if job["status"] in ("success", "failed"):
            return job
        time.sleep(0.3)
    raise TimeoutError(f"任务 {job_id} 轮询超时")


def download_report(client: TestClient, token: str, job_id: str, product: str, save_name: str):
    resp = client.get(f"/api/recon/jobs/{job_id}/download",
                      params={"product": product}, headers=auth(token))
    check(f"下载 {product} 报告返回200", resp.status_code == 200, resp.text)
    path = TMP_ROOT / save_name
    path.write_bytes(resp.content)
    return load_workbook(path).active


def main() -> int:
    print("=" * 70)
    print("M2 API 冒烟测试（基金估值价格核对，多产品批量）")
    print(f"临时目录: {TMP_ROOT}")
    print("=" * 70)

    with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
        expected_stats = json.load(f)

    with TestClient(app) as client:
        # 1. 用户准备
        print("--- 1. 用户准备 ---")
        admin_token = login_and_change(client, "admin", "Admin@123", "Admin@2026New")
        resp = client.post("/api/admin/users",
                           json={"username": "op04", "password": "Op@123456", "role": "operator"},
                           headers=auth(admin_token))
        check("创建 operator 返回200", resp.status_code == 200, resp.text)
        resp = client.post("/api/admin/users",
                           json={"username": "vw04", "password": "Vw@123456", "role": "viewer"},
                           headers=auth(admin_token))
        check("创建 viewer 返回200", resp.status_code == 200, resp.text)
        op_token = login_and_change(client, "op04", "Op@123456", "Op@654321New")
        vw_token = login_and_change(client, "vw04", "Vw@123456", "Vw@654321New")

        # 2. 科目取价规则种子核验
        print("--- 2. 规则种子核验 ---")
        resp = client.get("/api/admin/system/subject-price-rules", headers=auth(op_token))
        check("operator 查规则列表返回200", resp.status_code == 200, resp.text)
        rules = resp.json()
        check("种子规则为 2 条", len(rules) == 2, str(rules))
        seed = {r["subject_prefix"]: r["price_field"] for r in rules}
        check("种子含 1101→市价", seed.get("1101") == "市价", str(seed))
        check("种子含 1501→单位成本", seed.get("1501") == "单位成本", str(seed))
        resp = client.get("/api/admin/system/subject-price-rules", headers=auth(vw_token))
        check("viewer 查规则列表被拒(403)", resp.status_code == 403, resp.text)

        # 3. M2 基线任务全流程
        print("--- 3. M2 基线任务全流程 ---")
        job_id = create_m2_job(client, op_token)
        final = poll_job(client, op_token, job_id)
        check("M2 任务执行成功", final["status"] == "success",
              f"status={final['status']} error={final.get('error')}")
        check("统计摘要与 M2 黄金基线一致", final["stats"] == expected_stats,
              f"预期{expected_stats} 实际{final['stats']}")
        result_files = final.get("result_files") or []
        check("结果文件清单含 6301 报告", "6301_估值价格核对报告.xlsx" in result_files,
              str(result_files))
        check("结果文件清单含 6302 报告", "6302_估值价格核对报告.xlsx" in result_files,
              str(result_files))

        # 4. 按产品下载与内容抽样
        print("--- 4. 按产品下载与内容抽样 ---")
        ws = download_report(client, op_token, job_id, "6301", "r6301.xlsx")
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(REPORT_COLUMNS))]
        check("6301 报告表头==REPORT_COLUMNS", headers == REPORT_COLUMNS, str(headers))
        rgb = lambda r: ws.cell(row=r, column=1).fill.start_color.rgb or ""
        check("6301 第2行(110101)绿色一致", rgb(2).endswith("C6EFCE"), rgb(2))
        check("6301 第3行(110102)红色差异", rgb(3).endswith("FFC7CE"), rgb(3))
        check("6301 第5行(888880)橙色单边", rgb(5).endswith("FFC000"), rgb(5))
        check("6301 第8行(110108)橙色单边", rgb(8).endswith("FFC000"), rgb(8))
        check("6301 第4行(110103)估值表价格=100.0(取单位成本金丝雀)",
              float(ws.cell(row=4, column=6).value) == 100.0,
              str(ws.cell(row=4, column=6).value))
        check("6301 第4行备注含[摊余成本]",
              "摊余成本" in str(ws.cell(row=4, column=11).value),
              str(ws.cell(row=4, column=11).value))
        check("6301 第5行备注含[新标准券]",
              "新标准券" in str(ws.cell(row=5, column=11).value),
              str(ws.cell(row=5, column=11).value))
        summary = [ws.cell(row=10, column=c).value for c in range(1, 6)]
        check("6301 底部汇总行=7/2/2/3",
              summary == ["汇总", "总记录 7", "一致 2", "差异 2", "单边 3"], str(summary))

        ws2 = download_report(client, op_token, job_id, "6302", "r6302.xlsx")
        rgb2 = lambda r: ws2.cell(row=r, column=1).fill.start_color.rgb or ""
        check("6302 第2行(220201)绿色一致", rgb2(2).endswith("C6EFCE"), rgb2(2))
        check("6302 第3行(220202)红色差异", rgb2(3).endswith("FFC7CE"), rgb2(3))
        summary2 = [ws2.cell(row=5, column=c).value for c in range(1, 6)]
        check("6302 底部汇总行=2/1/1/0",
              summary2 == ["汇总", "总记录 2", "一致 1", "差异 1", "单边 0"], str(summary2))

        # 缺省下载 = 首个产品报告
        resp = client.get(f"/api/recon/jobs/{job_id}/download", headers=auth(op_token))
        check("缺省下载返回200", resp.status_code == 200, resp.text)
        check("缺省下载为 xlsx(PK头)", resp.content[:2] == b"PK", f"长度={len(resp.content)}")
        # 不存在的产品 → 404 中文
        resp = client.get(f"/api/recon/jobs/{job_id}/download",
                          params={"product": "9999"}, headers=auth(op_token))
        check("下载不存在产品返回404", resp.status_code == 404, resp.text)
        check("404 提示含产品标识", "9999" in resp.text and "不存在" in resp.text, resp.text)

        # 5. 配对异常 400
        print("--- 5. 配对异常 400 ---")
        resp = client.post("/api/recon/m2/jobs",
                           files=m2_files([SYS_6301, SYS_6302], [VAL_6301]),
                           headers=auth(op_token))
        check("落单配对返回400", resp.status_code == 400, resp.text)
        check("400 指明落单文件[6302]",
              "未能配对估值表" in resp.text and "6302" in resp.text, resp.text)

        no_id = TMP_ROOT / "新综合信息查询_基金证券.xlsx"
        no_id.write_bytes(SYS_6301.read_bytes())
        resp = client.post("/api/recon/m2/jobs",
                           files=m2_files([no_id], [VAL_6301]),
                           headers=auth(op_token))
        check("无产品标识返回400", resp.status_code == 400, resp.text)
        check("400 提示未识别产品标识", "未识别到产品标识" in resp.text, resp.text)

        bad_sys = TMP_ROOT / "新综合信息查询_基金证券-6301.xlsx"
        pd.DataFrame({"证券代码": ["110101"], "证券名称": ["测试"], "估值价格": [100.0]}
                     ).to_excel(bad_sys, index=False)
        resp = client.post("/api/recon/m2/jobs",
                           files=m2_files([bad_sys], [VAL_6301]),
                           headers=auth(op_token))
        check("系统端缺列返回400", resp.status_code == 400, resp.text)
        check("400 提示含缺失列名[持仓]", "持仓" in resp.text, resp.text)

        # 6. viewer 权限边界
        print("--- 6. viewer 权限边界 ---")
        resp = client.post("/api/recon/m2/jobs",
                           files=m2_files([SYS_6301], [VAL_6301]),
                           headers=auth(vw_token))
        check("viewer 建 M2 任务被拒(403)", resp.status_code == 403, resp.text)
        resp = client.get("/api/recon/jobs", params={"module": "M2"}, headers=auth(vw_token))
        check("viewer 可查 M2 历史(200)",
              resp.status_code == 200 and resp.json()["total"] == 1, resp.text)

        # 7. 规则 CRUD 权限与唯一性
        print("--- 7. 规则 CRUD 权限与唯一性 ---")
        resp = client.post("/api/admin/system/subject-price-rules",
                           json={"subject_prefix": "1102", "price_field": "市价"},
                           headers=auth(op_token))
        check("operator 新增规则被拒(403)", resp.status_code == 403, resp.text)
        resp = client.post("/api/admin/system/subject-price-rules",
                           json={"subject_prefix": "1101", "price_field": "成本"},
                           headers=auth(admin_token))
        check("重复前缀新增返回400", resp.status_code == 400, resp.text)
        check("400 提示前缀已存在", "已存在" in resp.text, resp.text)

        # 8. 热生效专项
        print("--- 8. 热生效专项（1102 规则即改即生效） ---")
        resp = client.post("/api/admin/system/subject-price-rules",
                           json={"subject_prefix": "1102", "price_field": "市价",
                                 "description": "其他投资", "sort_order": 3},
                           headers=auth(admin_token))
        check("admin 新增 1102 规则返回200", resp.status_code == 200, resp.text)
        rule_1102_id = resp.json()["id"]
        resp = client.put(f"/api/admin/system/subject-price-rules/{rule_1102_id}",
                          json={"note": "测试口径提示"}, headers=auth(admin_token))
        check("PUT 修改规则口径提示返回200",
              resp.status_code == 200 and resp.json()["note"] == "测试口径提示", resp.text)

        hot_job_id = create_m2_job(client, op_token)
        hot_final = poll_job(client, op_token, hot_job_id)
        check("热规则任务执行成功", hot_final["status"] == "success",
              f"error={hot_final.get('error')}")
        hot_6301 = (hot_final["stats"] or {}).get("products", {}).get("6301", {})
        check("热规则后 6301 统计=8/2/2/4",
              hot_6301 == {"总记录": 8, "一致": 2, "差异": 2, "单边": 4}, str(hot_6301))
        check("热规则后合计=10/3/3/4",
              (hot_final["stats"] or {}).get("合计")
              == {"总记录": 10, "一致": 3, "差异": 3, "单边": 4},
              str((hot_final["stats"] or {}).get("合计")))
        ws_hot = download_report(client, op_token, hot_job_id, "6301", "r6301_hot.xlsx")
        hot_labels = [ws_hot.cell(row=r, column=9).value for r in range(2, 10)]
        check("热规则报告含[1102-其他投资（取市价）]行",
              "1102-其他投资（取市价）" in hot_labels, str(hot_labels))
        row_1102 = hot_labels.index("1102-其他投资（取市价）") + 2 \
            if "1102-其他投资（取市价）" in hot_labels else None
        if row_1102:
            check("1102 行差异状态=估值表有系统无",
                  ws_hot.cell(row=row_1102, column=8).value == "估值表有系统无",
                  str(ws_hot.cell(row=row_1102, column=8).value))
        summary_hot = [ws_hot.cell(row=11, column=c).value for c in range(1, 6)]
        check("热规则 6301 底部汇总行=8/2/2/4",
              summary_hot == ["汇总", "总记录 8", "一致 2", "差异 2", "单边 4"],
              str(summary_hot))

        resp = client.delete(f"/api/admin/system/subject-price-rules/{rule_1102_id}",
                             headers=auth(admin_token))
        check("DELETE 1102 规则返回200", resp.status_code == 200, resp.text)
        back_job_id = create_m2_job(client, op_token)
        back_final = poll_job(client, op_token, back_job_id)
        check("删除规则后任务执行成功", back_final["status"] == "success",
              f"error={back_final.get('error')}")
        check("删除规则后统计回到黄金基线", back_final["stats"] == expected_stats,
              f"预期{expected_stats} 实际{back_final['stats']}")

        # 9. 审计日志落库核验
        print("--- 9. 审计日志 ---")
        conn = sqlite3.connect(os.environ["O32OPS_DB_PATH"])
        rows = dict(conn.execute(
            "SELECT action, COUNT(*) FROM sys_audit_log GROUP BY action").fetchall())
        for action in ("upload_create_job", "sys_subject_rule_create",
                       "sys_subject_rule_update", "sys_subject_rule_delete", "download"):
            check(f"审计动作已记录: {action}", rows.get(action, 0) > 0, str(rows))
        detail = conn.execute(
            "SELECT detail FROM sys_audit_log WHERE action='sys_subject_rule_create'").fetchone()
        check("规则新增审计含[1102→市价]",
              detail is not None and "1102→市价" in detail[0], str(detail))
        conn.close()

    print("=" * 70)
    if failures:
        print(f"M2 冒烟测试失败，共 {len(failures)} 项：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("M2 API 冒烟测试全部通过 ✅")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        sys.exit(2)
