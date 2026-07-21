# -*- coding: utf-8 -*-
"""
API 冒烟测试（纯脚本，fastapi.testclient + 临时目录数据/归档）

覆盖链路：
    1. admin 初始登录（Admin@123）→ 首登标记 → 未改密访问被拦截（403）
    2. 首登改密 → 创建 operator / viewer
    3. operator 登录 → 改密 → 上传黄金样本两表建 M1 任务 → 轮询至 success
    4. 统计摘要与黄金基线 expected_stats.json 逐项一致
    5. 同日期再次建任务 → 结果文件名自动 _v2（不覆盖）
    6. 下载结果 Excel → 抽样比对填充色（表头/大宗行/差异行/未匹配行）
    7. viewer 建任务被拒（403）、可查询历史（200）、下载被拒（403）
    8. 错误文件（列数不足 / 文件选反）返回 400 + 中文提示含"选反"
    9. 审计日志落库核验（login / user_create / upload_create_job / download）

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests\\api\\test_api_smoke.py
退出码：全部通过 0；任一失败非零。

作者：技术部
版本：1.0.0
日期：2026-07-17
"""

import json
import os
import sqlite3
import sys
import tempfile
import time
from pathlib import Path

# =============================================================================
# 环境准备：导入 app 之前注入测试配置（临时数据/归档目录、测试密钥）
# =============================================================================

SERVER_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = SERVER_ROOT.parent
TMP_ROOT = Path(tempfile.mkdtemp(prefix="o32ops_smoke_"))

os.environ["O32OPS_DATA_DIR"] = str(TMP_ROOT / "data")
os.environ["O32OPS_ARCHIVE_DIR"] = str(TMP_ROOT / "archive")
os.environ["O32OPS_DB_PATH"] = str(TMP_ROOT / "data" / "o32ops.db")
os.environ["O32OPS_SECRET_KEY"] = "smoke-test-secret-key-do-not-use-in-prod"

if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

import pandas as pd  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402

GOLDEN_DIR = SERVER_ROOT / "tests" / "golden"
FUND_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "基金资产表_样本.xlsx"
NETVALUE_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "净值查询表_样本.xlsx"
EXPECTED_STATS = GOLDEN_DIR / "expected" / "expected_stats.json"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login(client: TestClient, username: str, password: str) -> dict:
    resp = client.post("/api/auth/login", json={"username": username, "password": password})
    check(f"登录 {username} 返回200", resp.status_code == 200, resp.text)
    return resp.json()


def change_password(client: TestClient, token: str, old: str, new: str) -> None:
    resp = client.post("/api/auth/change-password",
                       json={"old_password": old, "new_password": new},
                       headers=auth_headers(token))
    check("修改密码返回200", resp.status_code == 200, resp.text)


def create_m1_job(client: TestClient, token: str) -> str:
    with open(FUND_SAMPLE, "rb") as f:
        fund_bytes = f.read()
    with open(NETVALUE_SAMPLE, "rb") as f:
        net_bytes = f.read()
    resp = client.post(
        "/api/recon/m1/jobs",
        files={
            "fund_file": (FUND_SAMPLE.name, fund_bytes, XLSX_MIME),
            "netvalue_file": (NETVALUE_SAMPLE.name, net_bytes, XLSX_MIME),
        },
        headers=auth_headers(token),
    )
    check("创建 M1 任务返回200", resp.status_code == 200, resp.text)
    return resp.json()["job_id"]


def poll_job(client: TestClient, token: str, job_id: str, timeout: float = 90.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/recon/jobs/{job_id}", headers=auth_headers(token))
        check(f"查询任务 {job_id} 返回200", resp.status_code == 200, resp.text)
        job = resp.json()
        if job["status"] in ("success", "failed"):
            return job
        time.sleep(0.3)
    raise TimeoutError(f"任务 {job_id} 轮询超时（{timeout}s）仍未结束")


def main() -> int:
    print("=" * 70)
    print("API 冒烟测试（O32 日常运维平台 一期 M1）")
    print(f"临时目录: {TMP_ROOT}")
    print("=" * 70)

    with TestClient(app) as client:
        # ------------------------------------------------------------------
        # 1. admin 首登与强制改密
        # ------------------------------------------------------------------
        print("--- 1. admin 首登 / 强制改密 ---")
        body = login(client, "admin", "Admin@123")
        check("admin 首登返回 must_change_password=True",
              body.get("must_change_password") is True, str(body))
        admin_token = body["access_token"]

        resp = client.get("/api/admin/users", headers=auth_headers(admin_token))
        check("未改密访问用户维护被拦截(403)", resp.status_code == 403, resp.text)
        check("拦截提示含'修改初始密码'", "修改初始密码" in resp.text, resp.text)

        change_password(client, admin_token, "Admin@123", "Admin@2026New")
        resp = client.get("/api/admin/users", headers=auth_headers(admin_token))
        check("改密后访问用户维护正常(200)", resp.status_code == 200, resp.text)
        check("初始用户列表仅 admin", [u["username"] for u in resp.json()] == ["admin"],
              resp.text)

        # ------------------------------------------------------------------
        # 2. 创建 operator / viewer
        # ------------------------------------------------------------------
        print("--- 2. 用户维护 ---")
        resp = client.post("/api/admin/users",
                           json={"username": "op01", "password": "Op@123456", "role": "operator",
                                 "display_name": "  张操作  ", "department": "运营部"},
                           headers=auth_headers(admin_token))
        check("创建 operator 返回200", resp.status_code == 200, resp.text)
        check("新用户强制首登改密", resp.json().get("must_change_password") is True, resp.text)
        check("新用户姓名/部门落库（去首尾空格）",
              resp.json().get("display_name") == "张操作"
              and resp.json().get("department") == "运营部",
              str(resp.json()))
        op01_id = resp.json()["id"]

        resp = client.post("/api/admin/users",
                           json={"username": "vw01", "password": "Vw@123456", "role": "viewer"},
                           headers=auth_headers(admin_token))
        check("创建 viewer 返回200", resp.status_code == 200, resp.text)
        check("选填字段缺省为 None",
              resp.json().get("display_name") is None
              and resp.json().get("department") is None, str(resp.json()))

        resp = client.post("/api/admin/users",
                           json={"username": "op01", "password": "Op@123456", "role": "operator"},
                           headers=auth_headers(admin_token))
        check("重复用户编号创建被拒(400)", resp.status_code == 400, resp.text)

        # display_name / department 修改与清空
        resp = client.put(f"/api/admin/users/{op01_id}",
                          json={"display_name": "张运营", "department": "  投资运营部  "},
                          headers=auth_headers(admin_token))
        check("修改姓名/部门返回200", resp.status_code == 200, resp.text)
        check("修改后姓名/部门已更新（去空格）",
              resp.json().get("display_name") == "张运营"
              and resp.json().get("department") == "投资运营部", str(resp.json()))
        resp = client.put(f"/api/admin/users/{op01_id}",
                          json={"department": ""}, headers=auth_headers(admin_token))
        check("空串清空部门（回退 None）", resp.status_code == 200
              and resp.json().get("department") is None
              and resp.json().get("display_name") == "张运营", resp.text)
        resp = client.get("/api/admin/users", headers=auth_headers(admin_token))
        op01_row = next(u for u in resp.json() if u["username"] == "op01")
        check("列表回读姓名/部门字段", op01_row["display_name"] == "张运营"
              and op01_row["department"] is None, str(op01_row))

        # ------------------------------------------------------------------
        # 3. operator 登录改密 → 建 M1 任务 → 轮询成功
        # ------------------------------------------------------------------
        print("--- 3. M1 核对任务全流程 ---")
        body = login(client, "op01", "Op@123456")
        op_token = body["access_token"]
        check("operator 首登需改密", body.get("must_change_password") is True, str(body))
        change_password(client, op_token, "Op@123456", "Op@654321New")

        job1 = create_m1_job(client, op_token)
        final1 = poll_job(client, op_token, job1)
        check("任务1 执行成功", final1["status"] == "success",
              f"status={final1['status']} error={final1.get('error')}")
        check("任务1 进度=100", final1["progress"] == 100, str(final1["progress"]))

        with open(EXPECTED_STATS, "r", encoding="utf-8") as f:
            expected_stats = json.load(f)
        check("任务1 统计摘要与黄金基线一致", final1["stats"] == expected_stats,
              f"预期{expected_stats} 实际{final1['stats']}")
        check("任务1 日志含中文步骤文案",
              any("核对完成" in line or "精确匹配完成" in line for line in final1["log_tail"]),
              str(final1["log_tail"][-5:]))

        # 同日期重复执行 → 版本号 _v2
        job2 = create_m1_job(client, op_token)
        final2 = poll_job(client, op_token, job2)
        check("任务2 执行成功", final2["status"] == "success",
              f"status={final2['status']} error={final2.get('error')}")
        check("同日期结果文件名自动 _v2",
              (final2.get("result_filename") or "").endswith("_v2.xlsx"),
              f"任务1={final1.get('result_filename')} 任务2={final2.get('result_filename')}")

        # ------------------------------------------------------------------
        # 4. 历史查询 + 下载 + 颜色抽样比对
        # ------------------------------------------------------------------
        print("--- 4. 历史查询 / 下载 / 颜色抽样 ---")
        resp = client.get("/api/recon/jobs", params={"module": "M1", "status": "success"},
                          headers=auth_headers(op_token))
        check("任务历史查询返回200", resp.status_code == 200, resp.text)
        check("历史查询返回2条成功任务", resp.json()["total"] == 2, resp.text)

        resp = client.get(f"/api/recon/jobs/{job2}/download", headers=auth_headers(op_token))
        check("下载结果返回200", resp.status_code == 200, resp.text)
        check("下载内容为 xlsx(PK头)", resp.content[:2] == b"PK", f"长度={len(resp.content)}")
        check("下载文件名带日期与版本", "filename" in resp.headers.get("content-disposition", ""),
              resp.headers.get("content-disposition", ""))

        download_path = TMP_ROOT / "downloaded.xlsx"
        download_path.write_bytes(resp.content)
        wb = load_workbook(download_path)
        ws = wb.active
        rgb = lambda r, c: ws.cell(row=r, column=c).fill.start_color.rgb or ""
        check("颜色抽样: 表头 B4C7DC", rgb(1, 1).endswith("B4C7DC"), rgb(1, 1))
        check("颜色抽样: 第2行大宗 FFC000", rgb(2, 1).endswith("FFC000"), rgb(2, 1))
        check("颜色抽样: 第3行差异 FFCCCC", rgb(3, 1).endswith("FFCCCC"), rgb(3, 1))
        check("颜色抽样: 第12行大宗 FFC000", rgb(12, 1).endswith("FFC000"), rgb(12, 1))
        check("颜色抽样: 第17行未匹配 FFFF99", rgb(17, 1).endswith("FFFF99"), rgb(17, 1))
        check("颜色抽样: 数据行数=17(含表头)", ws.max_row == 17, str(ws.max_row))

        # ------------------------------------------------------------------
        # 5. viewer 权限边界
        # ------------------------------------------------------------------
        print("--- 5. viewer 权限边界 ---")
        body = login(client, "vw01", "Vw@123456")
        vw_token = body["access_token"]
        change_password(client, vw_token, "Vw@123456", "Vw@654321New")

        with open(FUND_SAMPLE, "rb") as f:
            fund_bytes = f.read()
        with open(NETVALUE_SAMPLE, "rb") as f:
            net_bytes = f.read()
        resp = client.post(
            "/api/recon/m1/jobs",
            files={"fund_file": (FUND_SAMPLE.name, fund_bytes, XLSX_MIME),
                   "netvalue_file": (NETVALUE_SAMPLE.name, net_bytes, XLSX_MIME)},
            headers=auth_headers(vw_token),
        )
        check("viewer 建任务被拒(403)", resp.status_code == 403, resp.text)
        check("viewer 403 提示含'权限不足'", "权限不足" in resp.text, resp.text)

        resp = client.get("/api/recon/jobs", headers=auth_headers(vw_token))
        check("viewer 可查历史(200)", resp.status_code == 200, resp.text)
        resp = client.get(f"/api/recon/jobs/{job2}", headers=auth_headers(vw_token))
        check("viewer 可查任务详情与统计(200)",
              resp.status_code == 200 and resp.json()["stats"] == expected_stats, resp.text)
        resp = client.get(f"/api/recon/jobs/{job2}/download", headers=auth_headers(vw_token))
        check("viewer 下载被拒(403)", resp.status_code == 403, resp.text)

        # ------------------------------------------------------------------
        # 6. 错误输入：400 中文提示
        # ------------------------------------------------------------------
        print("--- 6. 错误输入处理 ---")
        bad_fund = TMP_ROOT / "坏基金资产表.xlsx"
        pd.DataFrame(columns=[f"列{i}" for i in range(10)]).to_excel(bad_fund, index=False)
        with open(bad_fund, "rb") as f:
            bad_bytes = f.read()
        with open(NETVALUE_SAMPLE, "rb") as f:
            net_bytes = f.read()
        resp = client.post(
            "/api/recon/m1/jobs",
            files={"fund_file": (bad_fund.name, bad_bytes, XLSX_MIME),
                   "netvalue_file": (NETVALUE_SAMPLE.name, net_bytes, XLSX_MIME)},
            headers=auth_headers(op_token),
        )
        check("基金表列数不足返回400", resp.status_code == 400, resp.text)
        check("400 提示含'列数不足'", "列数不足" in resp.text, resp.text)
        check("400 提示含'选反'引导语", "选反" in resp.text, resp.text)

        # 文件选反：fund 位置传 9 列净值表，net 位置传 28 列基金表
        with open(FUND_SAMPLE, "rb") as f:
            fund_bytes = f.read()
        resp = client.post(
            "/api/recon/m1/jobs",
            files={"fund_file": (NETVALUE_SAMPLE.name, net_bytes, XLSX_MIME),
                   "netvalue_file": (FUND_SAMPLE.name, fund_bytes, XLSX_MIME)},
            headers=auth_headers(op_token),
        )
        check("文件选反返回400", resp.status_code == 400, resp.text)
        check("选反 400 提示含'选反'引导语", "选反" in resp.text, resp.text)

    # ----------------------------------------------------------------------
    # 7. 审计日志落库核验（直接查库）
    # ----------------------------------------------------------------------
    print("--- 7. 审计日志 ---")
    conn = sqlite3.connect(os.environ["O32OPS_DB_PATH"])
    rows = dict(conn.execute(
        "SELECT action, COUNT(*) FROM sys_audit_log GROUP BY action").fetchall())
    conn.close()
    for action in ("login", "user_create", "upload_create_job", "download", "change_password"):
        check(f"审计动作已记录: {action}", rows.get(action, 0) > 0, str(rows))

    print("=" * 70)
    if failures:
        print(f"冒烟测试失败，共 {len(failures)} 项：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("API 冒烟测试全部通过 ✅")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        sys.exit(2)
