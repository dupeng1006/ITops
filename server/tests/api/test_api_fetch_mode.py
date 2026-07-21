# -*- coding: utf-8 -*-
"""
DS-F4 取数模式联动冒烟测试（纯脚本，fastapi.testclient + SQLite 内部测试方言）

核心证据（验收 V5）：
    SQLite 测试库表数据 = 黄金样本同数据 → 配查询模板 → db 模式跑 M1/M2
    → 统计与文件模式黄金基线**逐项一致**（M1: 16/14/1/1/2/1；M2: 9/3/3/3）
    → 结果 Excel 与黄金预期逐格一致（值 + 填充色），且与 file 模式结果逐格一致。

覆盖链路：
    1. admin 登录改密 → 创建 operator / viewer
    2. API 创建 sqlite 数据源（内部测试方言）与 6 个查询模板
       （显式列清单 + WHERE biz_date=:biz_date，验证参数绑定注入路径）
    3. 【V5】M1 db 模式全流程 + 基线一致 + 逐格一致 + 快照归档断言
       （CSV 快照 UTF-8-SIG 带 BOM、引擎消费件 xlsx、recon_result_file 索引、
       job.fetch_mode=db）
    4. M1 db 异常：模板不存在 404 / 模块不匹配 400 中文 / 缺 biz_date 400 /
       非法 fetch_mode 400 / viewer 403
    5. 【V5】M2 db 模式双产品全流程 + 基线一致 + 报告抽样 + 快照断言
       （估值表快照 3 行标题占位版式）；product 从模板名派生；bad JSON /
       重复 product / 模板不存在 / viewer 403
    6. 审计：db 取数任务审计 detail 含模板、数据源、biz_date

运行（工作目录 server/）：
    .venv\\Scripts\\python.exe tests/api/test_api_fetch_mode.py
退出码：全部通过 0；任一失败非零。

作者：技术部
版本：1.0.0
日期：2026-07-20
"""

import json
import os
import sqlite3
import sys
import tempfile
import time
from pathlib import Path

# =============================================================================
# 环境准备：导入 app 之前注入测试配置（临时数据/归档目录、测试密钥）
# =============================================================================

SERVER_ROOT = Path(__file__).resolve().parents[2]
TMP_ROOT = Path(tempfile.mkdtemp(prefix="o32ops_fetch_smoke_"))

os.environ["O32OPS_DATA_DIR"] = str(TMP_ROOT / "data")
os.environ["O32OPS_ARCHIVE_DIR"] = str(TMP_ROOT / "archive")
os.environ["O32OPS_DB_PATH"] = str(TMP_ROOT / "data" / "o32ops.db")
os.environ["O32OPS_SECRET_KEY"] = "fetch-smoke-test-secret-key-do-not-use-in-prod"

if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

import pandas as pd  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402

PROJECT_ROOT = SERVER_ROOT.parent
M1_FUND_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "基金资产表_样本.xlsx"
M1_NET_SAMPLE = PROJECT_ROOT / "samples" / "golden" / "净值查询表_样本.xlsx"
M1_EXPECTED_STATS = SERVER_ROOT / "tests" / "golden" / "expected" / "expected_stats.json"
M1_EXPECTED_EXCEL = SERVER_ROOT / "tests" / "golden" / "expected" / "核对结果_预期.xlsx"
M2_SAMPLE_DIR = SERVER_ROOT / "tests" / "golden" / "m2" / "samples"
M2_EXPECTED_STATS = SERVER_ROOT / "tests" / "golden" / "m2" / "expected" / "expected_stats.json"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
BIZ = "20260718"               # 测试业务日期（fixture 表 biz_date 列同值）
SQLITE_DB = TMP_ROOT / "test_source.db"

failures: list = []


def check(name: str, ok: bool, detail: str = "") -> None:
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {name}" + (f"  -- {detail}" if detail and not ok else ""))
    if not ok:
        failures.append(f"{name}: {detail}")


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def login_and_change(client: TestClient, username: str, old_pw: str, new_pw: str) -> str:
    resp = client.post("/api/auth/login", json={"username": username, "password": old_pw})
    check(f"登录 {username} 返回200", resp.status_code == 200, resp.text)
    token = resp.json()["access_token"]
    if resp.json().get("must_change_password"):
        resp2 = client.post("/api/auth/change-password",
                            json={"old_password": old_pw, "new_password": new_pw},
                            headers=auth(token))
        check(f"{username} 修改密码返回200", resp2.status_code == 200, resp2.text)
    return token


def poll_job(client: TestClient, token: str, job_id: str, timeout: float = 90.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = client.get(f"/api/recon/jobs/{job_id}", headers=auth(token))
        job = resp.json()
        if job["status"] in ("success", "failed"):
            return job
        time.sleep(0.3)
    raise TimeoutError(f"任务 {job_id} 轮询超时")


def compare_excel_cells(path_a: Path, path_b: Path, name: str) -> None:
    """逐格比对两个 Excel（值 + 填充色）"""
    wa = load_workbook(path_a).active
    wb = load_workbook(path_b).active
    check(f"{name} 表尺寸一致",
          (wa.max_row, wa.max_column) == (wb.max_row, wb.max_column),
          f"A=({wa.max_row},{wa.max_column}) B=({wb.max_row},{wb.max_column})")
    if (wa.max_row, wa.max_column) != (wb.max_row, wb.max_column):
        return
    mismatch = []
    for r in range(1, wa.max_row + 1):
        for c in range(1, wa.max_column + 1):
            va, vb = wa.cell(row=r, column=c).value, wb.cell(row=r, column=c).value
            if (va if va is not None else "") != (vb if vb is not None else ""):
                mismatch.append(f"({r},{c})值: {va!r} != {vb!r}")
            fa = wa.cell(row=r, column=c).fill.start_color.rgb or ""
            fb = wb.cell(row=r, column=c).fill.start_color.rgb or ""
            if fa != fb:
                mismatch.append(f"({r},{c})色: {fa} != {fb}")
    check(f"{name} 逐格值+填充色一致", not mismatch, "\n    ".join(mismatch[:15]))


def build_sqlite_fixture() -> None:
    """构造 SQLite 测试库：表数据 = 黄金样本同数据（+ biz_date 列供参数绑定）"""
    df_fund = pd.read_excel(M1_FUND_SAMPLE, header=0)
    df_net = pd.read_excel(M1_NET_SAMPLE, header=0)
    df_sys1 = pd.read_excel(M2_SAMPLE_DIR / "新综合信息查询_基金证券-6301.xlsx", header=0)
    df_sys2 = pd.read_excel(M2_SAMPLE_DIR / "新综合信息查询_基金证券-6302.xlsx", header=0)
    df_val1 = pd.read_excel(M2_SAMPLE_DIR / "证券投资基金估值表_6301-20260720.xlsx", skiprows=3, header=0)
    df_val2 = pd.read_excel(M2_SAMPLE_DIR / "证券投资基金估值表_6302-20260720.xlsx", skiprows=3, header=0)
    for df in (df_fund, df_net, df_sys1, df_sys2, df_val1, df_val2):
        df["biz_date"] = BIZ
    conn = sqlite3.connect(SQLITE_DB)
    try:
        df_fund.to_sql("t_fund", conn, index=False)
        df_net.to_sql("t_net", conn, index=False)
        df_sys1.to_sql("t_m2_sys_6301", conn, index=False)
        df_sys2.to_sql("t_m2_sys_6302", conn, index=False)
        df_val1.to_sql("t_m2_val_6301", conn, index=False)
        df_val2.to_sql("t_m2_val_6302", conn, index=False)
    finally:
        conn.close()
    print(f"SQLite 测试库已构造: {SQLITE_DB}（6 张表，数据=黄金样本）")


def main() -> int:
    print("=" * 70)
    print("DS-F4 取数模式联动冒烟测试（fetch_mode=db，SQLite 内部测试方言）")
    print(f"临时目录: {TMP_ROOT}")
    print("=" * 70)

    build_sqlite_fixture()
    expected_m1_stats = json.loads(M1_EXPECTED_STATS.read_text(encoding="utf-8"))
    expected_m2_stats = json.loads(M2_EXPECTED_STATS.read_text(encoding="utf-8"))

    with TestClient(app) as client:
        # 1. 用户准备
        print("--- 1. 用户准备 ---")
        admin_token = login_and_change(client, "admin", "Admin@123", "Admin@2026New")
        resp = client.post("/api/admin/users",
                           json={"username": "op05", "password": "Op@123456", "role": "operator"},
                           headers=auth(admin_token))
        check("创建 operator 返回200", resp.status_code == 200, resp.text)
        resp = client.post("/api/admin/users",
                           json={"username": "vw05", "password": "Vw@123456", "role": "viewer"},
                           headers=auth(admin_token))
        check("创建 viewer 返回200", resp.status_code == 200, resp.text)
        op_token = login_and_change(client, "op05", "Op@123456", "Op@654321New")
        vw_token = login_and_change(client, "vw05", "Vw@123456", "Vw@654321New")

        # 2. sqlite 数据源 + 6 查询模板
        print("--- 2. sqlite 数据源与查询模板 ---")
        resp = client.post("/api/datasources",
                           json={"name": "sqlite测试库", "db_type": "sqlite",
                                 "db_name": str(SQLITE_DB), "username": "na", "password": "na"},
                           headers=auth(admin_token))
        check("创建 sqlite 数据源返回200", resp.status_code == 200, resp.text)
        ds_id = resp.json()["id"]

        fund_cols = ", ".join(f"列{i}" for i in range(28))
        net_cols = ", ".join(f"列{i}" for i in range(9))
        val_cols = '"科目代码", "科目名称", "数量", "单位成本", "成本", "成本占净值%", "市价", "市值", "市值占净值%", "估值增值", "停牌信息"'
        params_def = {"biz_date": {"type": "date", "required": True, "label": "业务日期"}}
        tpl_defs = [
            ("基金资产查询模板", "m1_fund", f"SELECT {fund_cols} FROM t_fund WHERE biz_date = :biz_date"),
            ("净值查询模板", "m1_netvalue", f"SELECT {net_cols} FROM t_net WHERE biz_date = :biz_date"),
            ("M2系统端-6301", "m2_system",
             f"SELECT 证券代码, 证券名称, 持仓, 估值价格, 市值 FROM t_m2_sys_6301 WHERE biz_date = :biz_date"),
            ("M2估值表-6301", "m2_valuation", f"SELECT {val_cols} FROM t_m2_val_6301 WHERE biz_date = :biz_date"),
            ("M2系统端-6302", "m2_system",
             f"SELECT 证券代码, 证券名称, 持仓, 估值价格, 市值 FROM t_m2_sys_6302 WHERE biz_date = :biz_date"),
            ("M2估值表-6302", "m2_valuation", f"SELECT {val_cols} FROM t_m2_val_6302 WHERE biz_date = :biz_date"),
        ]
        tpl_ids = {}
        for name, module, sql in tpl_defs:
            resp = client.post("/api/query-templates",
                               json={"name": name, "module": module, "ds_id": ds_id,
                                     "sql_text": sql, "params_def": params_def},
                               headers=auth(admin_token))
            check(f"创建模板[{name}]返回200", resp.status_code == 200, resp.text)
            tpl_ids[name] = resp.json()["id"]

        # 3. 【V5】M1 db 模式全流程
        print("--- 3. V5：M1 db 模式全流程与基线一致 ---")
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "db",
                                 "fund_template_id": str(tpl_ids["基金资产查询模板"]),
                                 "netvalue_template_id": str(tpl_ids["净值查询模板"]),
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("M1 db 模式建任务返回200", resp.status_code == 200, resp.text)
        db_job_id = resp.json()["job_id"]
        db_job = poll_job(client, op_token, db_job_id)
        check("M1 db 任务执行成功", db_job["status"] == "success",
              f"error={db_job.get('error')}")
        check("【V5】db 模式统计==文件模式黄金基线(16/14/1/1/2/1)",
              db_job["stats"] == expected_m1_stats,
              f"预期{expected_m1_stats} 实际{db_job['stats']}")
        check("db 任务 fetch_mode=db", db_job.get("fetch_mode") == "db",
              str(db_job.get("fetch_mode")))

        # 下载 db 结果 Excel vs 黄金预期逐格
        resp = client.get(f"/api/recon/jobs/{db_job_id}/download", headers=auth(op_token))
        check("下载 db 模式 M1 结果返回200", resp.status_code == 200, resp.text)
        db_result = TMP_ROOT / "m1_db_result.xlsx"
        db_result.write_bytes(resp.content)
        compare_excel_cells(M1_EXPECTED_EXCEL, db_result, "【V5】db结果 vs 黄金预期Excel")

        # file 模式对照任务（同样本）→ 两结果逐格一致
        resp = client.post(
            "/api/recon/m1/jobs",
            files={"fund_file": (M1_FUND_SAMPLE.name, M1_FUND_SAMPLE.read_bytes(), XLSX_MIME),
                   "netvalue_file": (M1_NET_SAMPLE.name, M1_NET_SAMPLE.read_bytes(), XLSX_MIME)},
            data={"biz_date": BIZ},
            headers=auth(op_token))
        check("M1 file 模式建任务返回200", resp.status_code == 200, resp.text)
        file_job = poll_job(client, op_token, resp.json()["job_id"])
        check("M1 file 任务执行成功", file_job["status"] == "success",
              f"error={file_job.get('error')}")
        check("file 任务 fetch_mode=file", file_job.get("fetch_mode") == "file",
              str(file_job.get("fetch_mode")))
        resp = client.get(f"/api/recon/jobs/{file_job['job_id']}/download", headers=auth(op_token))
        file_result = TMP_ROOT / "m1_file_result.xlsx"
        file_result.write_bytes(resp.content)
        compare_excel_cells(file_result, db_result, "【V5】db结果 vs file结果")

        # 快照归档断言
        conn = sqlite3.connect(os.environ["O32OPS_DB_PATH"])
        input_files = [r[0] for r in conn.execute(
            "SELECT file_name FROM recon_result_file WHERE job_id=? AND file_type='input'",
            (db_job_id,)).fetchall()]
        conn.close()
        t1, t2 = tpl_ids["基金资产查询模板"], tpl_ids["净值查询模板"]
        check("输入索引含基金资产 CSV 快照", f"查询快照_基金资产表_tpl{t1}.csv" in input_files,
              str(input_files))
        check("输入索引含净值 CSV 快照", f"查询快照_净值查询表_tpl{t2}.csv" in input_files,
              str(input_files))
        check("输入索引含基金资产引擎消费件 xlsx",
              f"fund__查询快照_基金资产表_tpl{t1}.xlsx" in input_files, str(input_files))
        snap_dir = Path(os.environ["O32OPS_ARCHIVE_DIR"]) / "m1" / BIZ / db_job_id / "input"
        snap_csv = snap_dir / f"查询快照_基金资产表_tpl{t1}.csv"
        check("基金资产 CSV 快照已落归档", snap_csv.exists(), str(snap_dir))
        if snap_csv.exists():
            raw = snap_csv.read_bytes()
            check("CSV 快照为 UTF-8-SIG（带 BOM）", raw[:3] == b"\xef\xbb\xbf",
                  raw[:6].hex())
            snap_df = pd.read_csv(snap_csv)
            check("基金资产快照行数=15（与样本一致）", len(snap_df) == 15, str(len(snap_df)))
            check("基金资产快照列数=28", len(snap_df.columns) == 28, str(len(snap_df.columns)))
        snap_net_csv = snap_dir / f"查询快照_净值查询表_tpl{t2}.csv"
        if snap_net_csv.exists():
            check("净值快照行数=17", len(pd.read_csv(snap_net_csv)) == 17,
                  str(len(pd.read_csv(snap_net_csv))))

        # 4. M1 db 异常分支
        print("--- 4. M1 db 异常分支 ---")
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "db", "fund_template_id": "99999",
                                 "netvalue_template_id": str(t2), "biz_date": BIZ},
                           headers=auth(op_token))
        check("模板不存在返回404", resp.status_code == 404, resp.text)
        check("404 中文提示含模板 id", "查询模板不存在" in resp.text and "99999" in resp.text,
              resp.text)
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "db", "fund_template_id": str(t2),
                                 "netvalue_template_id": str(t2), "biz_date": BIZ},
                           headers=auth(op_token))
        check("模块不匹配返回400", resp.status_code == 400, resp.text)
        check("400 中文指明模块不匹配",
              "不能用作基金资产查询模板" in resp.text and "m1_netvalue" in resp.text, resp.text)
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "db", "fund_template_id": str(t1),
                                 "netvalue_template_id": str(t2)},
                           headers=auth(op_token))
        check("缺 biz_date 返回400", resp.status_code == 400, resp.text)
        check("400 提示必须提供 biz_date", "必须提供业务日期" in resp.text, resp.text)
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "ftp", "biz_date": BIZ},
                           headers=auth(op_token))
        check("非法 fetch_mode 返回400", resp.status_code == 400, resp.text)
        resp = client.post("/api/recon/m1/jobs",
                           data={"fetch_mode": "db", "fund_template_id": str(t1),
                                 "netvalue_template_id": str(t2), "biz_date": BIZ},
                           headers=auth(vw_token))
        check("viewer 建 db 任务被拒(403)", resp.status_code == 403, resp.text)

        # 5. 【V5】M2 db 模式双产品全流程
        print("--- 5. V5：M2 db 模式双产品全流程 ---")
        groups = [
            {"product": "6301", "system_template_id": tpl_ids["M2系统端-6301"],
             "valuation_template_id": tpl_ids["M2估值表-6301"]},
            {"product": "6302", "system_template_id": tpl_ids["M2系统端-6302"],
             "valuation_template_id": tpl_ids["M2估值表-6302"]},
        ]
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db", "groups_json": json.dumps(groups),
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("M2 db 模式建任务返回200", resp.status_code == 200, resp.text)
        m2_job_id = resp.json()["job_id"]
        m2_job = poll_job(client, op_token, m2_job_id)
        check("M2 db 任务执行成功", m2_job["status"] == "success",
              f"error={m2_job.get('error')}")
        check("【V5】M2 db 模式统计==黄金基线(9/3/3/3)",
              m2_job["stats"] == expected_m2_stats,
              f"预期{expected_m2_stats} 实际{m2_job['stats']}")

        resp = client.get(f"/api/recon/jobs/{m2_job_id}/download",
                          params={"product": "6301"}, headers=auth(op_token))
        check("下载 6301 报告返回200", resp.status_code == 200, resp.text)
        r6301 = TMP_ROOT / "m2_db_6301.xlsx"
        r6301.write_bytes(resp.content)
        ws = load_workbook(r6301).active
        rgb = lambda r: ws.cell(row=r, column=1).fill.start_color.rgb or ""
        check("6301 第2行(110101)绿色一致", rgb(2).endswith("C6EFCE"), rgb(2))
        check("6301 第3行(110102)红色差异", rgb(3).endswith("FFC7CE"), rgb(3))
        check("6301 第5行(888880)橙色单边", rgb(5).endswith("FFC000"), rgb(5))
        check("6301 第4行(110103)估值表价格=100.0(金丝雀)",
              float(ws.cell(row=4, column=6).value) == 100.0,
              str(ws.cell(row=4, column=6).value))
        summary = [ws.cell(row=10, column=c).value for c in range(1, 6)]
        check("6301 底部汇总行=7/2/2/3",
              summary == ["汇总", "总记录 7", "一致 2", "差异 2", "单边 3"], str(summary))

        # M2 快照断言
        m2_snap_dir = Path(os.environ["O32OPS_ARCHIVE_DIR"]) / "m2" / BIZ / m2_job_id / "input"
        t3, t4 = tpl_ids["M2系统端-6301"], tpl_ids["M2估值表-6301"]
        sys_snap = m2_snap_dir / f"system__6301__查询快照_系统端_tpl{t3}.csv"
        val_snap = m2_snap_dir / f"valuation__6301__查询快照_估值表_tpl{t4}.csv"
        check("M2 系统端快照已落归档", sys_snap.exists(), str(m2_snap_dir))
        check("M2 估值表快照已落归档", val_snap.exists(), str(m2_snap_dir))
        if sys_snap.exists():
            check("系统端快照带 BOM", sys_snap.read_bytes()[:3] == b"\xef\xbb\xbf")
            check("系统端快照行数=8（含零持仓与汇总行，引擎再清洗）",
                  len(pd.read_csv(sys_snap)) == 8, str(len(pd.read_csv(sys_snap))))
        if val_snap.exists():
            lines = val_snap.read_bytes().decode("utf-8-sig").splitlines()
            check("估值表快照第1行为标题占位", "数据库查询快照" in lines[0], lines[0])
            check("估值表快照第2行含模板与数据源",
                  "M2估值表-6301" in lines[1] and "sqlite测试库" in lines[1], lines[1])
            val_df = pd.read_csv(val_snap, skiprows=3)
            check("估值表快照 skiprows=3 后列数=11", len(val_df.columns) == 11,
                  str(list(val_df.columns)[:4]))
            check("估值表快照数据行=10", len(val_df) == 10, str(len(val_df)))

        # product 从模板名派生（不显式传 product）
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db",
                                 "groups_json": json.dumps([
                                     {"system_template_id": t3, "valuation_template_id": t4}]),
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("缺省 product 建任务返回200（模板名派生）", resp.status_code == 200, resp.text)
        derive_job = poll_job(client, op_token, resp.json()["job_id"])
        check("派生任务成功且产品=6301",
              derive_job["status"] == "success"
              and set((derive_job["stats"] or {}).get("products", {}).keys()) == {"6301"},
              str(derive_job.get("stats")))

        # M2 db 异常分支
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db", "groups_json": "{bad json",
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("groups_json 非法 JSON 返回400", resp.status_code == 400, resp.text)
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db",
                                 "groups_json": json.dumps([
                                     {"product": "6301", "system_template_id": t3,
                                      "valuation_template_id": t4},
                                     {"product": "6301", "system_template_id": t3,
                                      "valuation_template_id": t4}]),
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("重复 product 返回400", resp.status_code == 400, resp.text)
        check("400 指明重复产品标识", "重复产品标识 6301" in resp.text, resp.text)
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db",
                                 "groups_json": json.dumps([
                                     {"product": "6301", "system_template_id": 99999,
                                      "valuation_template_id": t4}]),
                                 "biz_date": BIZ},
                           headers=auth(op_token))
        check("M2 模板不存在返回404", resp.status_code == 404, resp.text)
        resp = client.post("/api/recon/m2/jobs",
                           data={"fetch_mode": "db", "groups_json": json.dumps(groups),
                                 "biz_date": BIZ},
                           headers=auth(vw_token))
        check("viewer 建 M2 db 任务被拒(403)", resp.status_code == 403, resp.text)

        # 6. 审计
        print("--- 6. 审计 ---")
        conn = sqlite3.connect(os.environ["O32OPS_DB_PATH"])
        rows = conn.execute(
            "SELECT detail FROM sys_audit_log WHERE action='upload_create_job' AND detail LIKE '%db 取数%'"
        ).fetchall()
        conn.close()
        details = " ".join(r[0] for r in rows)
        check("db 取数任务审计已落库", len(rows) >= 2, str(len(rows)))
        check("审计 detail 含模板名/数据源/biz_date",
              all(k in details for k in ("基金资产查询模板", "sqlite测试库", BIZ)), details[:200])
        check("审计 detail 含 M2 分组模板与数据源",
              all(k in details for k in ("M2系统端-6301", "M2估值表-6301")), details[:300])

    print("=" * 70)
    if failures:
        print(f"DS-F4 取数模式冒烟测试失败，共 {len(failures)} 项：")
        for item in failures:
            print(f"  - {item}")
        return 1
    print("DS-F4 取数模式冒烟测试全部通过 ✅")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        sys.exit(2)
