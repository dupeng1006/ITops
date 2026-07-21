# -*- coding: utf-8 -*-
"""
M3 引擎 —— 基金属性表银行间 ID 匹配（Interbank ID Matching）

业务规格（需求文档 V2.3 §3.4 / 既定 Skill）：
    1. 以 基金全称 ↔ 交易成员全称 做**精确匹配**（严禁模糊匹配/包含匹配；
       仅做首尾空白字符规整，不做任何相似度计算）；
    2. 匹配成功且新ID≠原ID → 「精确匹配-有变动」（绿 #C6EFCE）；
       相同 → 「精确匹配-无变化」（蓝 #BDD7EE）；
       未匹配 → 「未匹配」（红 #FFC7CE，保持原值）；
    3. 匹配成功的记录用新 交易成员ID 更新基金属性表的 银行间ID 字段；
    4. 类型陷阱：银行间ID 列可能因空值被 pandas 推断为 float64/int64，
       更新前必须统一转字符串（空值→''，整数值不带 .0），新 ID 统一为字符串；
    5. .XLS 老格式 OLE2 告警可容忍；严重解析错误时给出中文提示建议另存 .xlsx。

输出三件套（写入指定输出目录，文件名沿用 Skill 既有命名，不带日期）：
    - 基金属性_精确匹配更新.xlsx  （更新后的基金属性表，全列保留，三类颜色按行标注）
    - 精确匹配结果明细.xlsx       （基金全称/原银行间ID/新银行间ID/匹配类型，同色系按行标注）
    - 精确匹配说明.md             （统计 + 有变动明细 + 未匹配原因归类提示）

作者：技术部
版本：1.0.0
日期：2026-07-17
"""

import logging
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# 匹配类型
MATCH_CHANGED = "精确匹配-有变动"
MATCH_SAME = "精确匹配-无变化"
MATCH_NONE = "未匹配"

# 颜色（与 Skill 规格一致）
FILL_CHANGED = "C6EFCE"   # 绿：有变动
FILL_SAME = "BDD7EE"      # 蓝：无变化
FILL_NONE = "FFC7CE"      # 红：未匹配

# 输出文件名（沿用 Skill 既有命名，不带日期；归档按 jobId 目录隔离天然不覆盖）
OUTPUT_UPDATED_NAME = "基金属性_精确匹配更新.xlsx"
OUTPUT_DETAIL_NAME = "精确匹配结果明细.xlsx"
OUTPUT_NOTE_NAME = "精确匹配说明.md"

# 输入关键列
FUND_COL_NAME = "基金全称"
FUND_COL_ID = "银行间ID"
MEMBER_COL_NAME = "交易成员全称"
MEMBER_COL_ID = "交易成员ID"
MEMBER_COL_ORG = "机构代码"


def setup_engine_logger(name: str = "o32ops.m3") -> logging.Logger:
    """构造 M3 引擎默认日志记录器（中文日志，输出到 stdout）"""
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
    return logger


class M3InterbankIdEngine:
    """
    M3 基金属性表银行间 ID 匹配引擎（纯函数式，配置注入式接口与 M1 统一）

    使用示例：
        >>> engine = M3InterbankIdEngine()
        >>> result = engine.run("基金属性表.xlsx", "交易成员基本信息表.csv", "输出目录")
        >>> print(result["stats"])
    """

    MODULE_CODE = "M3"
    MODULE_NAME = "基金属性表银行间ID匹配"

    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger or setup_engine_logger()

    # ------------------------------------------------------------------
    # 输入读取与校验
    # ------------------------------------------------------------------

    def load_fund_table(self, fund_path) -> pd.DataFrame:
        """读取基金属性表（.xls/.xlsx；OLE2 告警容忍，解析失败给中文提示）"""
        self.logger.info(f"读取基金属性表: {fund_path}")
        try:
            with warnings.catch_warnings():
                # .XLS 老格式 OLE2/compdoc 告警白名单：降级为日志，不阻断
                warnings.simplefilter("ignore")
                df = pd.read_excel(fund_path, header=0)
        except ImportError as e:
            raise ValueError(
                f"读取 .XLS 老格式需要 xlrd 组件，当前环境未安装（{e}）。"
                f"请将文件另存为 .xlsx 后重试，或联系管理员安装 xlrd"
            )
        except Exception as e:  # noqa: BLE001
            raise ValueError(
                f"基金属性表解析失败: {e}。建议将文件另存为 .xlsx 格式后重新上传"
            )
        self.logger.info(f"基金属性表记录数: {len(df)}, 列数: {len(df.columns)}")
        return df

    def load_member_table(self, member_path) -> pd.DataFrame:
        """读取交易成员基本信息表（.CSV，GBK 编码）"""
        self.logger.info(f"读取交易成员基本信息表: {member_path}")
        try:
            df = pd.read_csv(member_path, encoding="gbk")
        except UnicodeDecodeError as e:
            raise ValueError(
                f"交易成员基本信息表编码解析失败（要求 GBK 编码）: {e}。"
                f"请确认 CSV 文件以 GBK/ANSI 编码保存"
            )
        except Exception as e:  # noqa: BLE001
            raise ValueError(f"交易成员基本信息表解析失败: {e}，请确认为 GBK 编码的 CSV 文件")
        self.logger.info(f"交易成员基本信息表记录数: {len(df)}, 列数: {len(df.columns)}")
        return df

    def validate_columns(self, df_fund: pd.DataFrame, df_member: pd.DataFrame) -> None:
        """
        关键列校验：缺列时抛中文 ValueError 并指明缺哪列

        Raises:
            ValueError: 基金属性表缺 基金全称/银行间ID；
                        交易成员表缺 交易成员全称/交易成员ID
        """
        fund_required = [FUND_COL_NAME, FUND_COL_ID]
        member_required = [MEMBER_COL_NAME, MEMBER_COL_ID]
        fund_missing = [c for c in fund_required if c not in df_fund.columns]
        member_missing = [c for c in member_required if c not in df_member.columns]
        if fund_missing:
            raise ValueError(
                f"基金属性表缺少关键列: {'、'.join(fund_missing)}。"
                f"请确认上传的是基金属性表（应包含 基金全称、银行间ID 列）"
            )
        if member_missing:
            raise ValueError(
                f"交易成员基本信息表缺少关键列: {'、'.join(member_missing)}。"
                f"请确认上传的是交易成员基本信息表（应包含 交易成员全称、交易成员ID、机构代码 列）"
            )

    def validate_input_files(self, fund_path, member_path) -> None:
        """读取并校验两个输入文件的关键列（供 API 层同步校验）"""
        df_fund = self.load_fund_table(fund_path)
        df_member = self.load_member_table(member_path)
        self.validate_columns(df_fund, df_member)
        self.logger.info("输入文件校验通过")

    # ------------------------------------------------------------------
    # 核心匹配逻辑
    # ------------------------------------------------------------------

    @staticmethod
    def normalize_id_series(series: pd.Series) -> pd.Series:
        """
        银行间ID 列字符串规整（类型陷阱处理）：
        - float64/int64（空值或数值存储导致）→ 空值为 ''，整数值不带 .0；
        - 其余 → astype(str)，'nan'/'NaN'/'None' 归一为 ''；
        返回值均为字符串，供比较与写回使用。
        """
        if pd.api.types.is_float_dtype(series) or pd.api.types.is_integer_dtype(series):
            def _num_to_str(v):
                if pd.isna(v):
                    return ""
                fv = float(v)
                return str(int(fv)) if fv.is_integer() else str(fv)
            return series.map(_num_to_str)
        result = series.astype(str)
        return result.replace({"nan": "", "NaN": "", "None": "", "NAType": ""})

    def exact_match(
        self, df_fund: pd.DataFrame, df_member: pd.DataFrame
    ) -> Tuple[pd.DataFrame, pd.Series, pd.Series]:
        """
        精确匹配（基金全称 ↔ 交易成员全称，仅首尾空白规整，严禁模糊/包含匹配）

        Returns:
            (df_updated, match_types, new_ids)
            df_updated: 更新银行间ID 后的基金属性表（全列保留，银行间ID 统一字符串）
            match_types: 每行匹配类型（有变动/无变化/未匹配）
            new_ids: 每行新银行间ID（未匹配行为 ''）
        """
        # 交易成员表：全称 → 交易成员ID（字符串）；重复全称取第一条并告警
        member = df_member.copy()
        member[MEMBER_COL_ID] = self.normalize_id_series(member[MEMBER_COL_ID])
        dup = member[MEMBER_COL_NAME].astype(str).str.strip().duplicated()
        if dup.any():
            dup_names = member.loc[dup, MEMBER_COL_NAME].astype(str).tolist()
            self.logger.warning(f"交易成员表存在重复全称（取第一条）: {dup_names}")
        name_to_id: Dict[str, str] = {}
        for _, row in member.iterrows():
            key = str(row[MEMBER_COL_NAME]).strip()
            if key not in name_to_id:
                name_to_id[key] = row[MEMBER_COL_ID]

        self.logger.info(f"交易成员全称索引构建完成: {len(name_to_id)} 条")

        # 基金属性表：原 银行间ID 规整为字符串
        df_updated = df_fund.copy()
        original_ids = self.normalize_id_series(df_updated[FUND_COL_ID])

        match_types: List[str] = []
        new_ids: List[str] = []
        for idx in range(len(df_updated)):
            fund_name = str(df_updated.iloc[idx][FUND_COL_NAME]).strip()
            original = original_ids.iloc[idx]
            if fund_name in name_to_id:
                new_id = name_to_id[fund_name]
                if new_id != original:
                    match_types.append(MATCH_CHANGED)
                else:
                    match_types.append(MATCH_SAME)
                new_ids.append(new_id)
            else:
                match_types.append(MATCH_NONE)
                new_ids.append("")  # 未匹配保持原值

        # 写回：匹配成功用新ID，未匹配保持原值（均为字符串，规避 float64 写入陷阱）
        df_updated[FUND_COL_ID] = [
            new_ids[i] if match_types[i] != MATCH_NONE else original_ids.iloc[i]
            for i in range(len(df_updated))
        ]

        changed = match_types.count(MATCH_CHANGED)
        same = match_types.count(MATCH_SAME)
        none = match_types.count(MATCH_NONE)
        self.logger.info(f"精确匹配完成: 有变动{changed}条, 无变化{same}条, 未匹配{none}条")
        return df_updated, pd.Series(match_types, name="匹配类型"), pd.Series(new_ids, name="新银行间ID")

    # ------------------------------------------------------------------
    # 未匹配原因归类（按名称特征自动归类，仅提示用途，不影响匹配结论）
    # ------------------------------------------------------------------

    def classify_unmatched(self, fund_name: str, member_names: List[str]) -> str:
        """
        未匹配原因归类提示：
        1. 名称以 TR_ 等内部前缀开头 → 内部简称；
        2. 与某交易成员全称互为包含关系 → 命名后缀不一致；
        3. 其余 → 未注册产品（交易成员表无对应记录）。
        """
        name = str(fund_name).strip()
        if name.upper().startswith("TR_"):
            return "内部简称（TR_ 前缀），请核对交易成员表中的对应全称"
        for member_name in member_names:
            if name and member_name and (name in member_name or member_name in name):
                return f"命名后缀不一致（交易成员表存在相近名称「{member_name}」），请人工确认"
        return "未注册产品（交易成员表无对应记录），请确认是否已完成注册"

    # ------------------------------------------------------------------
    # 输出三件套
    # ------------------------------------------------------------------

    @staticmethod
    def _fill_for(match_type: str) -> PatternFill:
        color = {
            MATCH_CHANGED: FILL_CHANGED,
            MATCH_SAME: FILL_SAME,
            MATCH_NONE: FILL_NONE,
        }[match_type]
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def write_updated_excel(
        self, df_updated: pd.DataFrame, match_types: pd.Series, output_path: Path
    ) -> None:
        """输出 1：更新后的基金属性表（全列保留，三类颜色按行标注）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "基金属性更新"
        header_font = Font(bold=True, size=11)

        for col_idx, col_name in enumerate(df_updated.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
        for row_idx in range(len(df_updated)):
            fill = self._fill_for(match_types.iloc[row_idx])
            for col_idx, col_name in enumerate(df_updated.columns, 1):
                value = df_updated.iloc[row_idx][col_name]
                cell = ws.cell(row=row_idx + 2, column=col_idx,
                               value=None if pd.isna(value) else value)
                cell.fill = fill
        wb.save(output_path)
        self.logger.info(f"输出文件已生成: {output_path}")

    def write_detail_excel(self, detail_df: pd.DataFrame, output_path: Path) -> None:
        """输出 2：匹配结果明细（基金全称/原银行间ID/新银行间ID/匹配类型）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "精确匹配明细"
        header_font = Font(bold=True, size=11)

        for col_idx, col_name in enumerate(detail_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
        for row_idx in range(len(detail_df)):
            fill = self._fill_for(detail_df.iloc[row_idx]["匹配类型"])
            for col_idx, col_name in enumerate(detail_df.columns, 1):
                value = detail_df.iloc[row_idx][col_name]
                cell = ws.cell(row=row_idx + 2, column=col_idx,
                               value=None if pd.isna(value) else value)
                cell.fill = fill
        wb.save(output_path)
        self.logger.info(f"输出文件已生成: {output_path}")

    def write_note_md(
        self,
        detail_df: pd.DataFrame,
        stats: Dict,
        unmatched_reasons: Dict[str, str],
        output_path: Path,
    ) -> None:
        """输出 3：匹配说明（统计 + 有变动明细 + 未匹配原因归类提示）"""
        lines: List[str] = []
        lines.append("# 基金属性表银行间ID 精确匹配说明")
        lines.append("")
        lines.append("## 一、统计摘要")
        lines.append("")
        for key, value in stats.items():
            lines.append(f"- {key}: {value}")
        lines.append("")
        lines.append("## 二、精确匹配-有变动明细")
        lines.append("")
        changed = detail_df[detail_df["匹配类型"] == MATCH_CHANGED]
        if len(changed) == 0:
            lines.append("（无）")
        else:
            lines.append("| 基金全称 | 原银行间ID | 新银行间ID |")
            lines.append("|---|---|---|")
            for _, row in changed.iterrows():
                lines.append(f"| {row['基金全称']} | {row['原银行间ID']} | {row['新银行间ID']} |")
        lines.append("")
        lines.append("## 三、未匹配记录与原因归类提示")
        lines.append("")
        unmatched = detail_df[detail_df["匹配类型"] == MATCH_NONE]
        if len(unmatched) == 0:
            lines.append("（无）")
        else:
            lines.append("| 基金全称 | 原银行间ID | 原因归类提示 |")
            lines.append("|---|---|---|")
            for _, row in unmatched.iterrows():
                reason = unmatched_reasons.get(str(row["基金全称"]).strip(), "待人工确认")
                lines.append(f"| {row['基金全称']} | {row['原银行间ID']} | {reason} |")
            lines.append("")
            lines.append("> 说明：原因归类为按名称特征的自动提示（内部简称如 TR_ 前缀、"
                         "未注册产品、命名后缀不一致），仅供参考，最终以人工确认为准。")
        lines.append("")
        output_path.write_text("\n".join(lines), encoding="utf-8")
        self.logger.info(f"输出文件已生成: {output_path}")

    # ------------------------------------------------------------------
    # 主流程
    # ------------------------------------------------------------------

    def run(
        self,
        fund_path,
        member_path,
        output_dir,
        rule_config=None,
    ) -> Dict:
        """
        执行 M3 银行间 ID 匹配

        Args:
            fund_path: 基金属性表路径（.xls/.xlsx）
            member_path: 交易成员基本信息表路径（.csv，GBK 编码）
            output_dir: 输出目录（三件套写入处）
            rule_config: 预留参数位（本模块无业务规则配置，仅为与 M1 引擎接口统一）

        Returns:
            dict: {
                "module": "M3",
                "stats": {总记录数, 精确匹配-有变动, 精确匹配-无变化, 未匹配},
                "detail_df": 匹配结果明细 DataFrame,
                "output_files": 输出文件路径列表 [更新表, 明细, 说明],
                "output_dir": 输出目录 str,
            }

        Raises:
            ValueError: 输入文件解析失败或缺少关键列（中文提示指明缺哪列）
        """
        self.logger.info("=" * 60)
        self.logger.info(f"开始 {self.MODULE_CODE} {self.MODULE_NAME}引擎执行")
        self.logger.info("=" * 60)

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # 1. 读取与校验
        df_fund = self.load_fund_table(fund_path)
        df_member = self.load_member_table(member_path)
        self.validate_columns(df_fund, df_member)

        # 2. 精确匹配与更新
        self.logger.info("正在执行精确匹配（基金全称 ↔ 交易成员全称）...")
        df_updated, match_types, new_ids = self.exact_match(df_fund, df_member)

        # 3. 明细 DataFrame
        original_ids = self.normalize_id_series(df_fund[FUND_COL_ID])
        detail_df = pd.DataFrame({
            "基金全称": df_fund[FUND_COL_NAME].astype(str),
            "原银行间ID": original_ids,
            "新银行间ID": [new_ids.iloc[i] if match_types.iloc[i] != MATCH_NONE else ""
                           for i in range(len(df_fund))],
            "匹配类型": match_types,
        })

        # 4. 统计
        stats = {
            "总记录数": int(len(df_fund)),
            MATCH_CHANGED: int((match_types == MATCH_CHANGED).sum()),
            MATCH_SAME: int((match_types == MATCH_SAME).sum()),
            MATCH_NONE: int((match_types == MATCH_NONE).sum()),
        }

        # 5. 未匹配原因归类
        member_names = [str(v).strip() for v in df_member[MEMBER_COL_NAME].tolist()]
        unmatched_reasons: Dict[str, str] = {}
        for i in range(len(df_fund)):
            if match_types.iloc[i] == MATCH_NONE:
                fund_name = str(df_fund.iloc[i][FUND_COL_NAME]).strip()
                unmatched_reasons[fund_name] = self.classify_unmatched(
                    df_fund.iloc[i][FUND_COL_NAME], member_names
                )

        # 6. 输出三件套
        updated_path = output_dir / OUTPUT_UPDATED_NAME
        detail_path = output_dir / OUTPUT_DETAIL_NAME
        note_path = output_dir / OUTPUT_NOTE_NAME
        self.write_updated_excel(df_updated, match_types, updated_path)
        self.write_detail_excel(detail_df, detail_path)
        self.write_note_md(detail_df, stats, unmatched_reasons, note_path)

        self.logger.info("=" * 60)
        self.logger.info("匹配完成")
        for key, value in stats.items():
            self.logger.info(f"  {key}: {value}")
        self.logger.info("=" * 60)

        return {
            "module": self.MODULE_CODE,
            "stats": stats,
            "detail_df": detail_df,
            "output_files": [str(updated_path), str(detail_path), str(note_path)],
            "output_dir": str(output_dir),
        }
