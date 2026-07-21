# -*- coding: utf-8 -*-
"""
M2 引擎 —— 基金估值价格核对（Valuation Price Reconciliation）

业务规格（需求文档 V2.3 §3.3 + Skill fund-valuation-price-reconciliation）：
    1. 输入（多产品批量，按产品分别出报告）：
       - 系统端：新综合信息查询_基金证券（第 1 行起含中文列头；关键列
         证券代码/证券名称/持仓/估值价格；过滤末尾无代码汇总行与持仓 0/空行）；
       - 财务端：证券投资基金估值表（skiprows=3，第 4 行为列头；科目层级
         结构，仅提取数量列可转数值的叶子节点，证券代码=科目代码后 6 位）。
    2. 科目取价规则（配置化，不固定 1101/1501）：按启用规则逐条提取
       （科目前缀过滤 → 叶子节点 → 按规则取价字段取价），规则来源于
       sys_subject_price_rule 表（本引擎仅接受注入，不感知存储）。
    3. 两步匹配：P1 证券代码精确 → P2 证券名称模糊（阈值 fuzzy_sim，
       与 M1 同一 SequenceMatcher 口径）。
    4. 差异判定（阈值 price_tol，默认 0.0001）：|系统价-估值表价| < tol 一致
       （绿 #C6EFCE）；≥ tol 差异（红 #FFC7CE）；单边（系统有估值表无 /
       估值表有系统无，橙 #FFC000）。
    5. 备注设计（设计选择记录）：
       - 已知正常单边（如新标准券 888880：质押回购过程科目，估值表通常
         不体现）的说明文案为**引擎常量** KNOWN_SINGLE_SIDE_NOTES——
         属于行业共识口径，不随租户规则变化，故不入规则表；
       - 摊余成本类科目（如 1501）差异属正常的提示语**随规则可配置**
         （sys_subject_price_rule.note），报告备注列自动带出；
       - 输出列在规格 10 列基础上追加第 11 列「备注」承载上述说明
         （规格列保持不变，备注为附加列，见 M2-F7 结果解读提示）。

输出（每个产品一份）：{产品标识}_估值价格核对报告.xlsx
    列：证券代码/证券名称/系统持仓/系统估值价格/估值表数量/估值表价格/
        价格差异/差异状态/科目类型/匹配方式/备注；
    底部汇总行：总记录/一致/差异/单边；颜色按判定标注。
    统计：按产品 总记录/一致/差异/单边 + 合计（写库 stats_json）。

作者：技术部
版本：1.0.0
日期：2026-07-20
"""

import logging
import sys
import warnings
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.engines.table_io import read_table

# 差异状态
STATUS_SAME = "一致"
STATUS_DIFF = "差异"
STATUS_ONLY_SYSTEM = "系统有估值表无"
STATUS_ONLY_VALUATION = "估值表有系统无"

# 颜色（与 Skill 规格一致）
FILL_SAME = "C6EFCE"        # 绿：一致
FILL_DIFF = "FFC7CE"        # 红：差异
FILL_SINGLE = "FFC000"      # 橙：单边

# 匹配方式
MATCH_CODE = "证券代码"
MATCH_NAME = "证券名称"
MATCH_NONE = "未匹配"

# 已知正常单边说明（引擎常量：行业共识口径，不随规则变化——设计选择见模块 docstring 第 5 点）
KNOWN_SINGLE_SIDE_NOTES = {
    "888880": "新标准券：质押回购业务过程科目，估值表通常不体现，属正常",
}

# 报告列（规格 10 列 + 备注附加列）
REPORT_COLUMNS = [
    "证券代码", "证券名称", "系统持仓", "系统估值价格",
    "估值表数量", "估值表价格", "价格差异", "差异状态",
    "科目类型", "匹配方式", "备注",
]

# 系统端关键列
SYS_COL_CODE = "证券代码"
SYS_COL_NAME = "证券名称"
SYS_COL_HOLD = "持仓"
SYS_COL_PRICE = "估值价格"

# 估值表关键列
VAL_COL_SUBJECT = "科目代码"
VAL_COL_SUBJECT_NAME = "科目名称"
VAL_COL_QTY = "数量"


@dataclass(frozen=True)
class SubjectPriceRuleConfig:
    """科目取价规则（配置注入载体；来源 sys_subject_price_rule 表）"""
    subject_prefix: str
    price_field: str
    description: str = ""
    note: str = ""
    sort_order: int = 0

    @property
    def subject_label(self) -> str:
        """科目类型标注（取自配置说明，如 1101-交易性金融资产（取市价））"""
        desc = self.description or "未命名科目"
        return f"{self.subject_prefix}-{desc}（取{self.price_field}）"


def setup_engine_logger(name: str = "o32ops.m2") -> logging.Logger:
    """构造 M2 引擎默认日志记录器（中文日志，输出到 stdout）"""
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
    return logger


class M2ValuationPriceEngine:
    """
    M2 基金估值价格核对引擎（配置注入式，与 M1/M3 引擎风格统一）

    使用示例：
        >>> rules = [SubjectPriceRuleConfig("1101", "市价", "交易性金融资产"),
        ...          SubjectPriceRuleConfig("1501", "单位成本", "债权投资")]
        >>> engine = M2ValuationPriceEngine(subject_rules=rules)
        >>> result = engine.run(
        ...     jobs=[{"product": "6301", "system_path": "...xls", "valuation_path": "...xls"}],
        ...     output_dir="输出目录")
        >>> print(result["stats"])
    """

    MODULE_CODE = "M2"
    MODULE_NAME = "基金估值价格核对"

    def __init__(
        self,
        subject_rules: Optional[List[SubjectPriceRuleConfig]] = None,
        fuzzy_sim: float = 0.5,
        price_tol: float = 0.0001,
        logger: Optional[logging.Logger] = None,
    ):
        """
        Args:
            subject_rules: 启用的科目取价规则（按 sort_order 排序注入）
            fuzzy_sim: 证券名称模糊匹配相似度阈值（沿用 rule_threshold.fuzzy_sim）
            price_tol: 价格差异容差（沿用 rule_threshold.price_tol，默认 0.0001）
            logger: 日志记录器
        """
        self.subject_rules = sorted(
            subject_rules or [], key=lambda r: (r.sort_order, r.subject_prefix))
        self.fuzzy_sim = fuzzy_sim
        self.price_tol = price_tol
        self.logger = logger or setup_engine_logger()

    # ------------------------------------------------------------------
    # 输入读取与校验
    # ------------------------------------------------------------------

    def load_system_file(self, path) -> pd.DataFrame:
        """读取系统端 新综合信息查询（第 1 行起含列头；支持查询快照 CSV；OLE2 告警容忍）"""
        self.logger.info(f"读取系统端文件: {path}")
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df = read_table(path, header=0)
        except ImportError as e:
            raise ValueError(
                f"读取 .XLS 老格式需要 xlrd 组件，当前环境未安装（{e}）。"
                f"请将文件另存为 .xlsx 后重试"
            )
        except Exception as e:  # noqa: BLE001
            raise ValueError(f"系统端文件解析失败: {e}。建议将文件另存为 .xlsx 格式后重新上传")
        required = [SYS_COL_CODE, SYS_COL_NAME, SYS_COL_HOLD, SYS_COL_PRICE]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise ValueError(
                f"系统端文件缺少关键列: {'、'.join(missing)}。"
                f"请确认上传的是 新综合信息查询_基金证券（应包含 证券代码、证券名称、持仓、估值价格 列）"
            )
        self.logger.info(f"系统端原始记录数: {len(df)}")
        return df

    def load_valuation_file(self, path, price_fields: Optional[List[str]] = None) -> pd.DataFrame:
        """读取财务端 证券投资基金估值表（skiprows=3，第 4 行为列头；支持查询快照 CSV）"""
        self.logger.info(f"读取估值表文件: {path}")
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df = read_table(path, skiprows=3, header=0)
        except ImportError as e:
            raise ValueError(
                f"读取 .XLS 老格式需要 xlrd 组件，当前环境未安装（{e}）。"
                f"请将文件另存为 .xlsx 后重试"
            )
        except Exception as e:  # noqa: BLE001
            raise ValueError(f"估值表文件解析失败: {e}。建议将文件另存为 .xlsx 格式后重新上传")
        required = [VAL_COL_SUBJECT, VAL_COL_SUBJECT_NAME, VAL_COL_QTY]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise ValueError(
                f"估值表缺少关键列: {'、'.join(missing)}。"
                f"请确认上传的是 证券投资基金估值表（第 4 行应为列头：科目代码、科目名称、数量、"
                f"单位成本、成本、成本占净值%、市价、市值、市值占净值%、估值增值、停牌信息）"
            )
        # 按启用规则校验取价字段列存在（缺列即配置与文件不匹配，提前中文报错）
        for field in (price_fields or []):
            if field not in df.columns:
                raise ValueError(
                    f"估值表缺少科目取价规则所需的取价字段列: {field}。"
                    f"请检查【系统管理-系统配置】中的科目取价规则与估值表列结构是否匹配"
                )
        self.logger.info(f"估值表原始记录数: {len(df)}")
        return df

    def validate_input_files(self, system_path, valuation_path) -> None:
        """读取并校验单产品两输入文件（供 API 层同步校验）"""
        self.load_system_file(system_path)
        price_fields = [r.price_field for r in self.subject_rules]
        self.load_valuation_file(valuation_path, price_fields=price_fields)
        self.logger.info("输入文件校验通过")

    # ------------------------------------------------------------------
    # 数据清洗与提取
    # ------------------------------------------------------------------

    @staticmethod
    def _to_numeric(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce")

    def clean_system_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """系统端清洗：过滤无代码汇总行与持仓 0/空行，价格/持仓转数值"""
        df = df.copy()
        df[SYS_COL_CODE] = df[SYS_COL_CODE].astype(str).str.strip()
        # 数值型证券代码兜底（如被 Excel 存为数值）：去 .0 后缀
        df[SYS_COL_CODE] = df[SYS_COL_CODE].str.replace(r"\.0$", "", regex=True)
        # 过滤汇总行（无证券代码 / nan 字符串 / 空串）
        df = df[df[SYS_COL_CODE].notna() & (df[SYS_COL_CODE] != "") & (df[SYS_COL_CODE].str.lower() != "nan")]
        df["_hold_num"] = self._to_numeric(df[SYS_COL_HOLD])
        # 过滤持仓为 0 或不可转数值的行
        df = df[df["_hold_num"].notna() & (df["_hold_num"] != 0)]
        df["_price_num"] = self._to_numeric(df[SYS_COL_PRICE])
        df[SYS_COL_NAME] = df[SYS_COL_NAME].astype(str).str.strip()
        filtered = len(df)
        self.logger.info(f"系统端清洗完成: 有效记录 {filtered} 条（已过滤汇总行与零持仓行）")
        return df

    def extract_valuation_detail(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        按启用的科目取价规则逐条提取估值表明细

        每条规则：科目前缀过滤 → 仅保留数量列可转数值的叶子节点 →
        证券代码=科目代码后 6 位 → 按规则取价字段取价 → 标注来源科目。
        """
        df = df.copy()
        df[VAL_COL_SUBJECT] = df[VAL_COL_SUBJECT].astype(str).str.strip()
        # 数值型科目代码兜底（如被 Excel 存为数值）：去 .0 后缀
        df[VAL_COL_SUBJECT] = df[VAL_COL_SUBJECT].str.replace(r"\.0$", "", regex=True)
        df["_qty_num"] = self._to_numeric(df[VAL_COL_QTY])

        parts: List[pd.DataFrame] = []
        for rule in self.subject_rules:
            mask = df[VAL_COL_SUBJECT].str.startswith(rule.subject_prefix)
            sub = df[mask & df["_qty_num"].notna()].copy()
            if len(sub) == 0:
                self.logger.info(f"科目 {rule.subject_prefix}（取{rule.price_field}）: 无叶子明细")
                continue
            sub["证券代码"] = sub[VAL_COL_SUBJECT].apply(
                lambda x: x[-6:] if len(x) >= 6 else x)
            sub["证券名称"] = sub[VAL_COL_SUBJECT_NAME].astype(str).str.strip()
            sub["估值表价格"] = self._to_numeric(sub[rule.price_field])
            sub["数量"] = sub["_qty_num"]
            sub["科目类型"] = rule.subject_label
            sub["_rule_note"] = rule.note or ""
            parts.append(sub[["证券代码", "证券名称", "数量", "估值表价格", "科目类型", "_rule_note"]])
            self.logger.info(
                f"科目 {rule.subject_prefix}（取{rule.price_field}）: 提取叶子明细 {len(sub)} 条")

        if not parts:
            return pd.DataFrame(
                columns=["证券代码", "证券名称", "数量", "估值表价格", "科目类型", "_rule_note"])
        result = pd.concat(parts, ignore_index=True)
        self.logger.info(f"估值表明细提取完成: 共 {len(result)} 条（{len(parts)} 条规则生效）")
        return result

    # ------------------------------------------------------------------
    # 两步匹配与差异判定
    # ------------------------------------------------------------------

    @staticmethod
    def _similarity(a: str, b: str) -> float:
        return SequenceMatcher(None, str(a), str(b)).ratio()

    def match_and_judge(self, sys_df: pd.DataFrame, val_df: pd.DataFrame) -> pd.DataFrame:
        """
        两步匹配（P1 证券代码精确 → P2 证券名称模糊）+ 差异判定

        Returns:
            报告 DataFrame（REPORT_COLUMNS 列结构，行序：系统侧记录在前、
            估值表单边在后，各自保持输入顺序）
        """
        records: List[dict] = []
        val_remaining = val_df.copy()  # 未被消费的估值表记录

        for _, srow in sys_df.iterrows():
            code = srow[SYS_COL_CODE]
            name = srow[SYS_COL_NAME]
            hold = srow["_hold_num"]
            price = srow["_price_num"]

            matched_idx = None
            match_way = MATCH_NONE

            # P1：证券代码精确匹配
            hit = val_remaining[val_remaining["证券代码"] == code]
            if len(hit) > 0:
                matched_idx = hit.index[0]
                match_way = MATCH_CODE
            else:
                # P2：证券名称模糊匹配（在剩余记录中找最佳相似度）
                best_score, best_idx = 0.0, None
                for vidx, vrow in val_remaining.iterrows():
                    score = self._similarity(name, vrow["证券名称"])
                    if score > best_score:
                        best_score, best_idx = score, vidx
                if best_idx is not None and best_score >= self.fuzzy_sim:
                    matched_idx = best_idx
                    match_way = MATCH_NAME
                    self.logger.info(
                        f"名称模糊匹配: {name} ↔ {val_remaining.loc[best_idx, '证券名称']}"
                        f"（相似度 {best_score:.3f}）"
                    )

            if matched_idx is not None:
                vrow = val_remaining.loc[matched_idx]
                val_remaining = val_remaining.drop(index=matched_idx)
                val_price = vrow["估值表价格"]
                diff = None if (pd.isna(price) or pd.isna(val_price)) else float(price) - float(val_price)
                if diff is None:
                    status = STATUS_DIFF  # 价格不可比（缺失）按差异处理并备注
                else:
                    status = STATUS_SAME if abs(diff) < self.price_tol else STATUS_DIFF
                note = ""
                if status == STATUS_DIFF and vrow["_rule_note"]:
                    note = vrow["_rule_note"]  # 口径提示（随规则配置，如 1501 摊余成本说明）
                records.append({
                    "证券代码": code, "证券名称": name,
                    "系统持仓": hold, "系统估值价格": price,
                    "估值表数量": vrow["数量"], "估值表价格": val_price,
                    "价格差异": diff, "差异状态": status,
                    "科目类型": vrow["科目类型"], "匹配方式": match_way,
                    "备注": note,
                })
            else:
                # 单边：系统有估值表无
                note = KNOWN_SINGLE_SIDE_NOTES.get(code, "")
                records.append({
                    "证券代码": code, "证券名称": name,
                    "系统持仓": hold, "系统估值价格": price,
                    "估值表数量": None, "估值表价格": None,
                    "价格差异": None, "差异状态": STATUS_ONLY_SYSTEM,
                    "科目类型": "", "匹配方式": MATCH_NONE,
                    "备注": note,
                })

        # 单边：估值表有系统无
        for _, vrow in val_remaining.iterrows():
            records.append({
                "证券代码": vrow["证券代码"], "证券名称": vrow["证券名称"],
                "系统持仓": None, "系统估值价格": None,
                "估值表数量": vrow["数量"], "估值表价格": vrow["估值表价格"],
                "价格差异": None, "差异状态": STATUS_ONLY_VALUATION,
                "科目类型": vrow["科目类型"], "匹配方式": MATCH_NONE,
                "备注": "",
            })

        return pd.DataFrame(records, columns=REPORT_COLUMNS)

    # ------------------------------------------------------------------
    # 报告输出（单产品一份 Excel）
    # ------------------------------------------------------------------

    @staticmethod
    def _fill_for(status: str) -> PatternFill:
        color = {
            STATUS_SAME: FILL_SAME,
            STATUS_DIFF: FILL_DIFF,
            STATUS_ONLY_SYSTEM: FILL_SINGLE,
            STATUS_ONLY_VALUATION: FILL_SINGLE,
        }[status]
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def write_report(self, report_df: pd.DataFrame, stats: Dict, output_path: Path) -> None:
        """生成单产品核对报告（颜色按判定；底部汇总行）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "估值价格核对"
        header_font = Font(bold=True, size=11)

        for col_idx, col_name in enumerate(REPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
        for row_idx in range(len(report_df)):
            row = report_df.iloc[row_idx]
            fill = self._fill_for(row["差异状态"])
            for col_idx, col_name in enumerate(REPORT_COLUMNS, 1):
                value = row[col_name]
                cell = ws.cell(row=row_idx + 2, column=col_idx,
                               value=None if (value is None or (isinstance(value, float) and pd.isna(value)))
                               else value)
                cell.fill = fill

        # 底部汇总行（空一行后写入）
        summary_row = len(report_df) + 3
        ws.cell(row=summary_row, column=1, value="汇总").font = header_font
        summary_items = [
            ("总记录", stats["总记录"]), ("一致", stats["一致"]),
            ("差异", stats["差异"]), ("单边", stats["单边"]),
        ]
        for offset, (label, count) in enumerate(summary_items, start=1):
            ws.cell(row=summary_row, column=1 + offset, value=f"{label} {count}")

        wb.save(output_path)
        self.logger.info(f"输出文件已生成: {output_path}")

    # ------------------------------------------------------------------
    # 主流程（多产品批量入口）
    # ------------------------------------------------------------------

    def run_product(self, product: str, system_path, valuation_path, output_dir: Path) -> Dict:
        """执行单产品核对：读入 → 清洗/提取 → 匹配判定 → 出报告"""
        self.logger.info("-" * 60)
        self.logger.info(f"产品 {product}: 开始核对")

        df_sys = self.clean_system_df(self.load_system_file(system_path))
        price_fields = [r.price_field for r in self.subject_rules]
        df_val = self.extract_valuation_detail(
            self.load_valuation_file(valuation_path, price_fields=price_fields))

        report_df = self.match_and_judge(df_sys, df_val)

        stats = {
            "总记录": int(len(report_df)),
            "一致": int((report_df["差异状态"] == STATUS_SAME).sum()),
            "差异": int((report_df["差异状态"] == STATUS_DIFF).sum()),
            "单边": int(report_df["差异状态"].isin([STATUS_ONLY_SYSTEM, STATUS_ONLY_VALUATION]).sum()),
        }
        output_path = output_dir / f"{product}_估值价格核对报告.xlsx"
        self.write_report(report_df, stats, output_path)

        self.logger.info(
            f"产品 {product} 核对完成: 总记录 {stats['总记录']}, 一致 {stats['一致']}, "
            f"差异 {stats['差异']}, 单边 {stats['单边']}"
        )
        return {
            "product": product,
            "stats": stats,
            "report_df": report_df,
            "output_file": str(output_path),
        }

    def run(self, jobs: List[Dict], output_dir) -> Dict:
        """
        执行 M2 多产品估值价格核对

        Args:
            jobs: [{"product": "6301", "system_path": ..., "valuation_path": ...}, ...]
            output_dir: 输出目录（各产品报告写入处）

        Returns:
            dict: {
                "module": "M2",
                "stats": {"products": {product: {总记录,一致,差异,单边}}, "合计": {...}},
                "results": [单产品结果 dict 列表],
                "output_files": 输出文件路径列表,
                "output_dir": 输出目录 str,
            }

        Raises:
            ValueError: 输入文件解析失败、缺关键列或未配置科目取价规则（中文提示）
        """
        self.logger.info("=" * 60)
        self.logger.info(f"开始 {self.MODULE_CODE} {self.MODULE_NAME}引擎执行（{len(jobs)} 个产品）")
        self.logger.info("=" * 60)

        if not self.subject_rules:
            raise ValueError(
                "未配置任何启用的科目取价规则，请先在【系统管理-系统配置】中维护规则"
                "（默认应预置 1101→市价、1501→单位成本）"
            )
        self.logger.info(
            "科目取价规则: "
            + "；".join(f"{r.subject_prefix}→{r.price_field}" for r in self.subject_rules)
            + f"；fuzzy_sim={self.fuzzy_sim}，price_tol={self.price_tol}"
        )

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = []
        for job in jobs:
            results.append(self.run_product(
                job["product"], job["system_path"], job["valuation_path"], output_dir))

        product_stats = {r["product"]: r["stats"] for r in results}
        totals = {
            "总记录": sum(s["总记录"] for s in product_stats.values()),
            "一致": sum(s["一致"] for s in product_stats.values()),
            "差异": sum(s["差异"] for s in product_stats.values()),
            "单边": sum(s["单边"] for s in product_stats.values()),
        }
        stats = {"products": product_stats, "合计": totals}

        self.logger.info("=" * 60)
        self.logger.info(f"全部产品核对完成（合计: {totals}）")
        self.logger.info("=" * 60)

        return {
            "module": self.MODULE_CODE,
            "stats": stats,
            "results": results,
            "output_files": [r["output_file"] for r in results],
            "output_dir": str(output_dir),
        }
