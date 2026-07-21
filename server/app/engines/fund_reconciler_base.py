#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# 【基准代码复制件 · 逻辑零改动】
# 来源：samples/reference/fund_reconciler.py（已验证的核对小程序，验收基准）
# 版本：v1.0（基准原文标注版本 1.0.0，日期 2026-07-14）
# 说明：本文件为 M1 引擎封装的依赖基准，除本注释头外与原文件逐字节一致，
#       任何逻辑修改必须先升级 samples/reference 基准并重新验证。
# =============================================================================
"""
基金资产与净值核对工具 (Fund Asset & Net Value Reconciliation Tool)

功能：
    以信托计划代码为精确匹配键，辅以组合名称与产品名称模糊匹配，
    对比基金资产表与多账套净值查询表中的总资产与资产净值差异，
    生成带颜色标注的 Excel 核对结果。

合规要求：
    - 支持重命名映射规则配置
    - 支持映射去重（保留映射记录，删除原始记录）
    - 支持 I9 记录过滤
    - 支持差异阈值配置
    - 支持特殊产品（原大宗产品）独立标注（差异说明/行填充色可配）
    - 生成符合内部汇报格式的 Excel 报告

作者：技术部
版本：1.0.0
日期：2026-07-14
"""

import os
import sys
import argparse
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
import numpy as np
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# 配置常量
# =============================================================================

# 默认重命名映射规则（净值查询表信托计划代码按估值表规则替换）
DEFAULT_RENAME_MAP = {
    '7001': '6001',
    '7002': '7001',
    'AZ0001': '6101',
    'AZ0003': '6103',
    'AZ0004': '6104',
    'AZ0006': '6106',
    'AZ0007': '6107',
    'AZ0008': '6108',
    'AZ0010': '6110',
    'AZ0011': '6111',
    'AZ0012': '6112',
    'AZ0015': '6115',
    'AZ0016': '6116',
    'AZ0017': '6117',
    'AZ0018': '6118',
    'AZ0019': '6119',
    'AZ0020': '6120',
    'AZ0022': '6122',
    'AZ0023': '6123',
    'AZ0024': '6124',
    'AZ0026': '6126',
}

# 默认特殊产品清单（原大宗产品；标橙色，备注"大宗产品无需核对"）
# 值：{"note": 差异说明(None=默认文案), "color": 行填充色(6位HEX)}
DEFAULT_SPECIAL_PRODUCTS = {
    'AZ0206': {'note': None, 'color': 'FFC000'},
    'AZ0205': {'note': None, 'color': 'FFC000'},
    'AZ0221': {'note': None, 'color': 'FFC000'},
    'AZ0230': {'note': None, 'color': 'FFC000'},
    'AZ0130': {'note': None, 'color': 'FFC000'},
    'AZ0160': {'note': None, 'color': 'FFC000'},
    'AZ0208': {'note': None, 'color': 'FFC000'},
    'AZ0236': {'note': None, 'color': 'FFC000'},
    'AZ0210': {'note': None, 'color': 'FFC000'},
    'AZ0151': {'note': None, 'color': 'FFC000'},
    'AZ0207': {'note': None, 'color': 'FFC000'},
}

# 特殊产品默认行填充色
DEFAULT_SPECIAL_COLOR = 'FFC000'

# 默认差异阈值（%）
DEFAULT_DIFF_THRESHOLD = 1.0

# 默认模糊匹配相似度阈值
DEFAULT_SIMILARITY_THRESHOLD = 0.5

# 日志格式
LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'


# =============================================================================
# 工具函数
# =============================================================================

def similarity(a: str, b: str) -> float:
    """计算两个字符串的相似度（0-1）"""
    return SequenceMatcher(None, str(a), str(b)).ratio()


def setup_logging(level: int = logging.INFO) -> logging.Logger:
    """配置日志"""
    logging.basicConfig(
        level=level,
        format=LOG_FORMAT,
        handlers=[
            logging.StreamHandler(sys.stdout),
        ]
    )
    return logging.getLogger(__name__)


# =============================================================================
# 核心核对类
# =============================================================================

class FundAssetReconciler:
    """
    基金资产与净值核对器
    
    使用示例：
        >>> reconciler = FundAssetReconciler()
        >>> result = reconciler.reconcile(
        ...     fund_path='基金资产表.xls',
        ...     netvalue_path='净值查询表.xls',
        ...     output_path='核对结果.xlsx'
        ... )
    """
    
    def __init__(
        self,
        rename_map: Optional[Dict[str, str]] = None,
        special_products: Optional[Dict] = None,
        diff_threshold: float = DEFAULT_DIFF_THRESHOLD,
        similarity_threshold: float = DEFAULT_SIMILARITY_THRESHOLD,
        logger: Optional[logging.Logger] = None,
    ):
        """
        初始化核对器
        
        Args:
            rename_map: 净值查询表信托计划代码重命名映射
            special_products: 特殊产品配置 {代码: {"note"/"color"} 或 规则对象}；
                向后兼容：传入 set/list（旧大宗代码集合）时按默认
                note=None/color=FFC000 展开
            diff_threshold: 差异阈值（%）
            similarity_threshold: 模糊匹配相似度阈值
            logger: 日志记录器
        """
        self.rename_map = rename_map or DEFAULT_RENAME_MAP.copy()
        self.special_products = self._normalize_special_products(special_products)
        self.diff_threshold = diff_threshold
        self.similarity_threshold = similarity_threshold
        self.logger = logger or setup_logging()
        
        # 核对结果数据
        self.result_df: Optional[pd.DataFrame] = None
        self.stats: Optional[Dict] = None
    
    @staticmethod
    def _normalize_special_products(special_products: Optional[Dict]) -> Dict[str, Dict]:
        """归一化特殊产品配置为 {代码: {"note": ..., "color": ...}}（向后兼容 set/list）"""
        if special_products is None:
            return {k: dict(v) for k, v in DEFAULT_SPECIAL_PRODUCTS.items()}
        if isinstance(special_products, (set, frozenset, list, tuple)):
            return {str(code): {'note': None, 'color': DEFAULT_SPECIAL_COLOR}
                    for code in special_products}
        normalized = {}
        for code, spec in dict(special_products).items():
            if spec is None:
                normalized[str(code)] = {'note': None, 'color': DEFAULT_SPECIAL_COLOR}
            elif isinstance(spec, dict):
                normalized[str(code)] = {
                    'note': spec.get('note') or None,
                    'color': (spec.get('color') or DEFAULT_SPECIAL_COLOR).upper().lstrip('#'),
                }
            else:  # 规则对象（note/color 属性，如 SpecialProductRule）
                normalized[str(code)] = {
                    'note': getattr(spec, 'note', None) or None,
                    'color': (getattr(spec, 'color', None) or DEFAULT_SPECIAL_COLOR)
                             .upper().lstrip('#'),
                }
        return normalized
    
    def load_data(
        self,
        fund_path: Union[str, Path],
        netvalue_path: Union[str, Path],
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        读取基金资产表和净值查询表
        
        Args:
            fund_path: 基金资产表路径
            netvalue_path: 净值查询表路径
            
        Returns:
            (df_fund, df_net) 两个DataFrame
        """
        self.logger.info(f"读取基金资产表: {fund_path}")
        df_fund = pd.read_excel(fund_path, header=0)
        self.logger.info(f"基金资产表记录数: {len(df_fund)}")
        
        self.logger.info(f"读取净值查询表: {netvalue_path}")
        df_net = pd.read_excel(netvalue_path, header=0)
        self.logger.info(f"净值查询表记录数: {len(df_net)}")
        
        return df_fund, df_net
    
    def extract_key_columns(self, df_fund: pd.DataFrame, df_net: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        提取关键列
        
        基金资产表关键列：
            - 组合编号 (Col 2)
            - 组合名称 (Col 3)
            - 基金资产净值 (Col 11)
            - 基金总资产 (Col 27)
            - 信托计划代码 (Col 19)
            
        净值查询表关键列：
            - 信托计划代码 (Col 8)
            - 查询总资产 (Col 4)
            - 查询资产净值 (Col 2)
            - 产品名称 (Col 0)
        """
        # 基金资产表
        df_fund_key = df_fund.iloc[:, [2, 3, 11, 27, 19]].copy()
        df_fund_key.columns = ['组合编号', '组合名称', '基金资产净值', '基金总资产', '信托计划代码']
        
        # 净值查询表
        df_net_key = df_net.iloc[:, [8, 4, 2, 0]].copy()
        df_net_key.columns = ['信托计划代码', '查询总资产', '查询资产净值', '产品名称']
        
        self.logger.info(f"提取关键列完成: 基金资产表{len(df_fund_key)}行, 净值查询表{len(df_net_key)}行")
        
        return df_fund_key, df_net_key
    
    def apply_rename_mapping(self, df_net: pd.DataFrame) -> pd.DataFrame:
        """
        应用重命名映射规则
        
        1. 保存原始代码
        2. 应用映射规则
        3. 标记映射记录
        """
        df = df_net.copy()
        df['原始代码'] = df['信托计划代码'].astype(str)
        df['信托计划代码'] = df['原始代码'].replace(self.rename_map)
        df['映射记录'] = df['原始代码'] != df['信托计划代码']
        
        mapped_count = df['映射记录'].sum()
        self.logger.info(f"映射规则应用完成: {mapped_count}条记录被映射")
        
        return df
    
    def deduplicate_mapping(self, df_net: pd.DataFrame) -> pd.DataFrame:
        """
        映射去重
        
        规则：同一信托计划代码下，若同时存在原始记录和映射记录，
              删除原始记录，保留映射记录
        """
        grouped = df_net.groupby('信托计划代码', sort=False)
        filtered_rows = []
        
        for code, group in grouped:
            if group['映射记录'].any() and not group['映射记录'].all():
                # 同时存在映射记录和原始记录：仅保留映射记录
                filtered_rows.append(group[group['映射记录']])
                self.logger.debug(f"去重: 信托计划代码{code} 删除原始记录，保留映射记录")
            else:
                # 无映射记录或全部为映射记录：全部保留
                filtered_rows.append(group)
        
        df = pd.concat(filtered_rows, ignore_index=True).drop(columns=['原始代码', '映射记录'])
        self.logger.info(f"映射去重完成: {len(df)}条记录")
        
        return df
    
    def filter_i9_records(self, df_net: pd.DataFrame) -> pd.DataFrame:
        """
        I9 记录过滤
        
        规则：
        - 同一信托计划代码下有多条记录，且存在非I9记录 → 仅保留非I9记录
        - 同一信托计划代码下只有1条记录 → 保留
        - 同一信托计划代码下全部含I9 → 全部保留
        """
        df = df_net.copy()
        df['含I9'] = df['产品名称'].str.contains('I9', case=False, na=False)
        
        grouped = df.groupby('信托计划代码', sort=False)
        filtered_rows = []
        
        for code, group in grouped:
            if len(group) > 1 and not group['含I9'].all():
                filtered_rows.append(group[~group['含I9']])
                self.logger.debug(f"I9过滤: 信托计划代码{code} 过滤I9记录")
            else:
                filtered_rows.append(group)
        
        df = pd.concat(filtered_rows, ignore_index=True).drop(columns=['含I9'])
        self.logger.info(f"I9过滤完成: {len(df)}条记录")
        
        return df
    
    def exact_match(
        self,
        df_fund: pd.DataFrame,
        df_net: pd.DataFrame,
    ) -> pd.DataFrame:
        """
        精确匹配（以信托计划代码为Key）
        """
        merged = pd.merge(df_fund, df_net, on='信托计划代码', how='left')
        merged['匹配方式'] = np.where(merged['查询总资产'].notna(), '精确', '未匹配')
        merged['匹配相似度'] = None
        
        exact_count = (merged['匹配方式'] == '精确').sum()
        self.logger.info(f"精确匹配完成: {exact_count}条记录匹配成功")
        
        return merged
    
    def fuzzy_match(self, df_merged: pd.DataFrame, df_net: pd.DataFrame) -> pd.DataFrame:
        """
        模糊匹配（组合名称 vs 产品名称）
        
        对未匹配记录，使用SequenceMatcher计算相似度，
        相似度 >= 阈值则标记为模糊匹配。
        """
        unmatched = df_merged[df_merged['查询总资产'].isna()].copy()
        matched_codes = set(df_merged[df_merged['查询总资产'].notna()]['信托计划代码'].dropna().unique())
        available_net = df_net[~df_net['信托计划代码'].isin(matched_codes)].copy()
        
        used_net_indices = set()
        fuzzy_count = 0
        
        for idx1, row1 in unmatched.iterrows():
            best_score = 0
            best_idx2 = None
            best_row2 = None
            
            for idx2, row2 in available_net.iterrows():
                if idx2 in used_net_indices:
                    continue
                score = similarity(str(row1['组合名称']), str(row2['产品名称']))
                if score > best_score:
                    best_score = score
                    best_idx2 = idx2
                    best_row2 = row2
            
            if best_score >= self.similarity_threshold and best_idx2 is not None:
                df_merged.loc[idx1, '查询总资产'] = best_row2['查询总资产']
                df_merged.loc[idx1, '查询资产净值'] = best_row2['查询资产净值']
                df_merged.loc[idx1, '产品名称'] = best_row2['产品名称']
                df_merged.loc[idx1, '匹配方式'] = '模糊'
                df_merged.loc[idx1, '匹配相似度'] = round(best_score, 4)
                used_net_indices.add(best_idx2)
                fuzzy_count += 1
        
        self.logger.info(f"模糊匹配完成: {fuzzy_count}条记录匹配成功")
        
        return df_merged
    
    def calculate_differences(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        计算差异金额和差异百分比
        """
        df['总资产差值'] = df['基金总资产'] - df['查询总资产']
        df['总资产差值百分比'] = np.where(
            df['查询总资产'].notna() & (df['查询总资产'] != 0),
            (df['基金总资产'] - df['查询总资产']) / df['查询总资产'] * 100,
            np.nan
        )
        
        df['资产净值差值'] = df['基金资产净值'] - df['查询资产净值']
        df['资产净值差值百分比'] = np.where(
            df['查询资产净值'].notna() & (df['查询资产净值'] != 0),
            (df['基金资产净值'] - df['查询资产净值']) / df['查询资产净值'] * 100,
            np.nan
        )
        
        return df
    
    def sort_results(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        排序：未匹配放最后，其余按最大差异百分比降序
        """
        df['最大差异百分比'] = df[['总资产差值百分比', '资产净值差值百分比']].abs().max(axis=1)
        df['未匹配'] = df['查询总资产'].isna()
        df = df.sort_values(by=['未匹配', '最大差异百分比'], ascending=[True, False])
        df = df.drop(columns=['最大差异百分比', '未匹配'])
        df = df.reset_index(drop=True)
        
        return df
    
    def mark_bulk_products(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        标记特殊产品（内部列名"是否大宗"保持不变，避免黄金回归变动）
        """
        df['是否大宗'] = df['信托计划代码'].isin(self.special_products.keys())
        return df
    
    def generate_excel(self, df: pd.DataFrame, output_path: Union[str, Path]) -> None:
        """
        生成带颜色标注的Excel报告
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "核对结果"
        
        # 颜色定义（特殊产品行填充色按产品 color 缓存生成，默认 FFC000）
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        header_fill = PatternFill(start_color='B4C7DC', end_color='B4C7DC', fill_type='solid')
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        special_fills: Dict[str, PatternFill] = {}

        def special_fill_of(product_code: str) -> PatternFill:
            color = self.special_products.get(product_code, {}).get('color') or DEFAULT_SPECIAL_COLOR
            if color not in special_fills:
                special_fills[color] = PatternFill(
                    start_color=color, end_color=color, fill_type='solid')
            return special_fills[color]
        
        # 表头
        headers = [
            '组合编号', '组合名称', '信托计划代码', '产品名称',
            '匹配方式', '匹配相似度',
            '基金总资产', '查询总资产', '总资产差值', '总资产差值百分比(%)',
            '基金资产净值', '查询资产净值', '资产净值差值', '资产净值差值百分比(%)',
            '差异原因'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = header_fill
        
        # 数据行
        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2
            
            is_bulk = row.get('是否大宗', False)
            is_unmatched = row['匹配方式'] == '未匹配'
            is_diff = False
            
            if not is_unmatched and not is_bulk:
                total_diff = abs(row['总资产差值百分比']) if pd.notna(row['总资产差值百分比']) else 0
                nav_diff = abs(row['资产净值差值百分比']) if pd.notna(row['资产净值差值百分比']) else 0
                is_diff = total_diff > self.diff_threshold or nav_diff > self.diff_threshold
            
            # 差异原因（特殊产品：自定义差异说明，缺省默认文案）
            reason = ''
            if is_bulk:
                note = self.special_products.get(row['信托计划代码'], {}).get('note')
                reason = note or '大宗产品无需核对'
            
            values = [
                row['组合编号'], row['组合名称'], row['信托计划代码'], row['产品名称'],
                row['匹配方式'], row['匹配相似度'],
                row['基金总资产'], row['查询总资产'], row['总资产差值'], row['总资产差值百分比'],
                row['基金资产净值'], row['查询资产净值'], row['资产净值差值'], row['资产净值差值百分比'],
                reason
            ]
            
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=col, value=val)
                cell.border = thin_border
                
                if col in [7, 8, 9, 11, 12, 13]:
                    cell.number_format = '#,##0.00'
                elif col in [10, 14] and val is not None:
                    cell.number_format = '0.00'
                
                if is_bulk:
                    cell.fill = special_fill_of(row['信托计划代码'])
                elif is_unmatched:
                    cell.fill = yellow_fill
                elif is_diff:
                    cell.fill = red_fill
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 22
        ws.column_dimensions['O'].width = 20
        
        ws.freeze_panes = 'A2'
        wb.save(output_path)
        self.logger.info(f"Excel报告已生成: {output_path}")
    
    def calculate_statistics(self, df: pd.DataFrame) -> Dict:
        """
        计算核对统计指标
        """
        total = len(df)
        exact = (df['匹配方式'] == '精确').sum()
        fuzzy = (df['匹配方式'] == '模糊').sum()
        unmatched = (df['匹配方式'] == '未匹配').sum()
        bulk_count = df['是否大宗'].sum()
        
        # 非大宗且已匹配的记录
        diff_df = df[(df['匹配方式'] != '未匹配') & (~df['是否大宗'])].copy()
        diff_df['最大差异百分比'] = diff_df[['总资产差值百分比', '资产净值差值百分比']].abs().max(axis=1)
        diff_count = (diff_df['最大差异百分比'] > self.diff_threshold).sum()
        
        stats = {
            '总记录数': int(total),
            '精确匹配': int(exact),
            '模糊匹配': int(fuzzy),
            '未匹配': int(unmatched),
            '大宗产品数': int(bulk_count),
            f'差异>{self.diff_threshold}%数量（非大宗）': int(diff_count),
        }
        
        return stats
    
    def reconcile(
        self,
        fund_path: Union[str, Path],
        netvalue_path: Union[str, Path],
        output_path: Optional[Union[str, Path]] = None,
    ) -> Dict:
        """
        执行完整核对流程
        
        Args:
            fund_path: 基金资产表路径
            netvalue_path: 净值查询表路径
            output_path: 输出Excel路径（可选）
            
        Returns:
            核对统计信息字典
        """
        self.logger.info("=" * 60)
        self.logger.info("开始基金资产与净值核对")
        self.logger.info("=" * 60)
        
        # 1. 读取数据
        df_fund, df_net = self.load_data(fund_path, netvalue_path)
        
        # 2. 提取关键列
        df_fund_key, df_net_key = self.extract_key_columns(df_fund, df_net)
        
        # 3. 应用重命名映射
        df_net_mapped = self.apply_rename_mapping(df_net_key)
        
        # 4. 映射去重
        df_net_deduped = self.deduplicate_mapping(df_net_mapped)
        
        # 5. I9过滤
        df_net_filtered = self.filter_i9_records(df_net_deduped)
        
        # 6. 精确匹配
        df_merged = self.exact_match(df_fund_key, df_net_filtered)
        
        # 7. 模糊匹配
        df_merged = self.fuzzy_match(df_merged, df_net_filtered)
        
        # 8. 计算差异
        df_merged = self.calculate_differences(df_merged)
        
        # 9. 排序
        df_merged = self.sort_results(df_merged)
        
        # 10. 标记大宗产品
        df_merged = self.mark_bulk_products(df_merged)
        
        # 11. 保存结果
        self.result_df = df_merged
        
        # 12. 生成Excel
        if output_path:
            self.generate_excel(df_merged, output_path)
        
        # 13. 统计
        self.stats = self.calculate_statistics(df_merged)
        
        self.logger.info("=" * 60)
        self.logger.info("核对完成")
        for key, value in self.stats.items():
            self.logger.info(f"  {key}: {value}")
        self.logger.info("=" * 60)
        
        return self.stats


# =============================================================================
# 命令行接口
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='基金资产与净值核对工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python fund_reconciler.py fund.xls netvalue.xls -o result.xlsx
  python fund_reconciler.py fund.xls netvalue.xls --threshold 0.5 --similarity 0.6
        """
    )
    
    parser.add_argument('fund', help='基金资产表路径 (.xls/.xlsx)')
    parser.add_argument('netvalue', help='多账套净值查询表路径 (.xls/.xlsx)')
    parser.add_argument('-o', '--output', help='输出Excel路径')
    parser.add_argument('-t', '--threshold', type=float, default=DEFAULT_DIFF_THRESHOLD,
                        help=f'差异阈值 (%%), 默认 {DEFAULT_DIFF_THRESHOLD}')
    parser.add_argument('-s', '--similarity', type=float, default=DEFAULT_SIMILARITY_THRESHOLD,
                        help=f'模糊匹配相似度阈值, 默认 {DEFAULT_SIMILARITY_THRESHOLD}')
    parser.add_argument('-c', '--config', help='JSON配置文件路径（包含rename_map和special_products/bulk_products）')
    parser.add_argument('-v', '--verbose', action='store_true', help='显示详细日志')
    
    args = parser.parse_args()
    
    # 配置日志
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logger = setup_logging(log_level)
    
    # 加载配置（兼容新格式 special_products 对象数组与旧格式 bulk_products 字符串数组）
    rename_map = DEFAULT_RENAME_MAP.copy()
    special_products = {k: dict(v) for k, v in DEFAULT_SPECIAL_PRODUCTS.items()}
    
    if args.config:
        with open(args.config, 'r', encoding='utf-8') as f:
            config = json.load(f)
            rename_map.update(config.get('rename_map', {}))
            for item in config.get('special_products', []) or []:
                special_products[str(item['code'])] = {
                    'note': item.get('note') or None,
                    'color': (item.get('color') or DEFAULT_SPECIAL_COLOR).upper().lstrip('#'),
                }
            for code in config.get('bulk_products', []) or []:
                special_products.setdefault(
                    str(code), {'note': None, 'color': DEFAULT_SPECIAL_COLOR})
        logger.info(f"已加载配置文件: {args.config}")
    
    # 确定输出路径
    if not args.output:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.output = f'基金资产与净值核对结果_{timestamp}.xlsx'
    
    # 执行核对
    reconciler = FundAssetReconciler(
        rename_map=rename_map,
        special_products=special_products,
        diff_threshold=args.threshold,
        similarity_threshold=args.similarity,
        logger=logger,
    )
    
    stats = reconciler.reconcile(
        fund_path=args.fund,
        netvalue_path=args.netvalue,
        output_path=args.output,
    )
    
    print("\n核对结果:")
    for key, value in stats.items():
        print(f"  {key}: {value}")
    print(f"\n输出文件: {args.output}")


if __name__ == '__main__':
    main()
